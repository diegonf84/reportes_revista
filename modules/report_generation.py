"""Validation and publication controls for report generation."""

from __future__ import annotations

import logging
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook

from modules.common import validate_period


SPECIAL_COMPANY_CODES = (541, 686, 829)
CORRECTED_TABLES = (
    "base_subramos_corregida_actual",
    "base_ramos_corregida_actual",
    "base_cias_corregida_actual",
)
HISTORICAL_TABLES = ("base_subramos", "base_ramos")


@dataclass(frozen=True)
class CsvContract:
    separator: str
    columns: tuple[str, ...]


CSV_CONTRACTS: Mapping[str, CsvContract] = {
    "cuadro_nuevo": CsvContract(
        ",",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "primas_emitidas",
            "disponibilidades", "inversiones", "inmuebles", "deudas_total_aseg",
            "deudas_con_asegurados_ac_reaseguros", "deudas_neto",
            "patrimonio_neto", "inmuebles_inversion", "inmuebles_uso_propio",
        ),
    ),
    "cuadro_principal": CsvContract(
        ";",
        (
            "ramo_denominacion", "nombre_corto", "cod_cia", "primas_emitidas",
            "primas_devengadas", "siniestros", "pct_stros", "gastos",
            "pct_gastos", "resultado", "pct_result",
        ),
    ),
    "ganaron_perdieron": CsvContract(
        ";",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "resultado_tecnico", "pct_rt",
            "resultado_financiero", "pct_rf", "resultado_operaciones",
            "impuesto_ganancias", "resultado", "pct_result", "primas_devengadas",
        ),
    ),
    "apertura_por_subramo": CsvContract(
        ";",
        (
            "subramo_denominacion", "nombre_corto", "cod_cia", "primas",
            "porcentaje", "porcentaje_acumulado",
        ),
    ),
    "apertura_por_subramo_comparativo": CsvContract(
        ";",
        (
            "subramo_denominacion", "nombre_corto", "cod_cia", "primas",
            "primas_anterior", "variacion", "porcentaje", "porcentaje_acumulado",
        ),
    ),
    "primas_cedidas_reaseguro": CsvContract(
        ";",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "primas_emitidas",
            "primas_cedidas", "pct_cesion", "primas_retenidas", "pct_ret",
        ),
    ),
    "ranking_comparativo": CsvContract(
        ";",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "primas_emitidas",
            "variacion", "primas_anterior",
        ),
    ),
    "ranking_comparativo_por_ramo": CsvContract(
        ";",
        (
            "ramo_denominacion", "nombre_corto", "cod_cia", "primas_emitidas",
            "variacion", "primas_anterior",
        ),
    ),
    "sueldos_y_gastos": CsvContract(
        ";",
        (
            "nombre_corto", "cod_cia", "tipo_cia", "total_primas_devengadas",
            "total_gs_prod", "total_sueldos", "total_gs", "gs_reaseguro",
        ),
    ),
    "detalle_inmuebles": CsvContract(
        ",",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "inmuebles_inversion",
            "inmuebles_uso_propio", "inmuebles_total",
        ),
    ),
    "detalle_gastos": CsvContract(
        ",",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "primas_emitidas",
            "total_primas_devengadas", "total_gs_prod", "total_gs_explot",
            "total_gs", "pct_gastos_produccion", "pct_gastos_explotacion",
            "pct_gastos_totales",
        ),
    ),
    "distribucion_inversiones": CsvContract(
        ",",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "total_inv", "total_inv_inm",
            "total_inv_liq", "pct_inmuebles_inversion", "pct_inversiones_liquidas",
        ),
    ),
    "indicadores_solvencia": CsvContract(
        ",",
        (
            "tipo_cia", "nombre_corto", "cod_cia", "inmuebles_inversion",
            "inversiones", "disponibilidades", "deudas_con_asegurados",
        ),
    ),
}

EXCEL_REPORT_NAMES = (
    "apertura_por_subramo",
    "apertura_por_subramo_comparativo",
    "cuadro_nuevo",
    "cuadro_principal",
    "detalle_gastos",
    "detalle_inmuebles",
    "distribucion_inversiones",
    "ganaron_perdieron",
    "indicadores_solvencia",
    "primas_cedidas_reaseguro",
    "ranking_comparativo",
    "ranking_comparativo_por_ramo",
    "ranking_produccion",
    "sueldos_y_gastos",
)


class ReportValidationError(ValueError):
    """Raised when report inputs or staged outputs violate their contract."""


class ReportGenerationFailure(RuntimeError):
    """Functional report-generation failure safe to expose through the UI."""

    def __init__(
        self,
        message: str,
        *,
        status: str = "failed",
        logs: list[str] | None = None,
        failed_csv_reports: list[str] | None = None,
        failed_excel_reports: list[str] | None = None,
    ) -> None:
        super().__init__(message)
        self.status = status
        self.logs = logs or []
        self.failed_csv_reports = failed_csv_reports or []
        self.failed_excel_reports = failed_excel_reports or []


def expected_historical_periods(period: int) -> set[int]:
    """Return the periods required by the corrected-table calculations."""
    validate_period(period)
    year = period // 100
    quarter = period % 100
    previous_same_quarter = int(f"{year - 1}{quarter:02d}")

    if quarter == 1:
        return {
            period,
            previous_same_quarter,
            int(f"{year - 1}02"),
            int(f"{year - 1}04"),
            int(f"{year - 2}02"),
            int(f"{year - 2}04"),
        }
    if quarter == 2:
        return {
            period,
            int(f"{year - 1}02"),
            int(f"{year - 1}04"),
            int(f"{year - 2}02"),
            int(f"{year - 2}04"),
        }
    return {
        period,
        previous_same_quarter,
        int(f"{year}02"),
        int(f"{year - 1}02"),
    }


def _connect_read_only(database_path: str | Path) -> sqlite3.Connection:
    path = Path(database_path).expanduser().resolve()
    if not path.is_file():
        raise ReportValidationError("La base de datos configurada no existe.")
    return sqlite3.connect(f"file:{path}?mode=ro", uri=True)


def _table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def _periods_in_table(connection: sqlite3.Connection, table_name: str) -> set[int]:
    return {
        int(row[0])
        for row in connection.execute(
            f'SELECT DISTINCT periodo FROM "{table_name}" WHERE periodo IS NOT NULL'
        ).fetchall()
    }


def validate_report_preflight(database_path: str | Path, period: int) -> None:
    """Validate that all period-bound report inputs are current and complete."""
    validate_period(period)
    required_history = expected_historical_periods(period)

    with _connect_read_only(database_path) as connection:
        current_period_tables = (*CORRECTED_TABLES, "base_otros_conceptos")
        for table_name in current_period_tables:
            if not _table_exists(connection, table_name):
                raise ReportValidationError(
                    f"Falta la tabla requerida {table_name}. Procese las tablas antes de generar reportes."
                )

            columns = {
                row[1]
                for row in connection.execute(
                    f'PRAGMA table_info("{table_name}")'
                ).fetchall()
            }
            if "periodo" not in columns:
                raise ReportValidationError(
                    f"La tabla {table_name} no permite verificar su período."
                )

            table_periods = _periods_in_table(connection, table_name)
            if table_periods != {period}:
                displayed = ", ".join(map(str, sorted(table_periods))) or "sin datos"
                raise ReportValidationError(
                    f"La tabla {table_name} no corresponde exclusivamente al período {period} "
                    f"(contiene: {displayed})."
                )

            row_count = connection.execute(
                f'SELECT COUNT(*) FROM "{table_name}"'
            ).fetchone()[0]
            if row_count == 0:
                raise ReportValidationError(f"La tabla {table_name} está vacía.")

        for table_name in HISTORICAL_TABLES:
            if not _table_exists(connection, table_name):
                raise ReportValidationError(f"Falta la tabla histórica requerida {table_name}.")

            available_periods = _periods_in_table(connection, table_name)
            missing_periods = sorted(required_history - available_periods)
            if missing_periods:
                missing = ", ".join(map(str, missing_periods))
                raise ReportValidationError(
                    f"La tabla {table_name} no tiene la cobertura histórica requerida: {missing}."
                )

            placeholders = ",".join("?" for _ in SPECIAL_COMPANY_CODES)
            for required_period in sorted(required_history):
                present_codes = {
                    int(row[0])
                    for row in connection.execute(
                        f"""
                        SELECT DISTINCT CAST(cod_cia AS INTEGER)
                        FROM "{table_name}"
                        WHERE periodo = ?
                          AND CAST(cod_cia AS INTEGER) IN ({placeholders})
                        """,
                        (required_period, *SPECIAL_COMPANY_CODES),
                    ).fetchall()
                }
                missing_codes = sorted(set(SPECIAL_COMPANY_CODES) - present_codes)
                if missing_codes:
                    codes = ", ".join(f"{code:04d}" for code in missing_codes)
                    raise ReportValidationError(
                        f"Faltan compañías especiales en {table_name} para {required_period}: {codes}."
                    )


def expected_csv_filenames(period: str) -> set[str]:
    return {f"{period}_{name}.csv" for name in CSV_CONTRACTS}


def expected_excel_filenames(period: str) -> set[str]:
    return {f"{period}_{name}.xlsx" for name in EXCEL_REPORT_NAMES}


def validate_csv_outputs(directory: str | Path, period: str) -> list[Path]:
    """Validate the exact staged CSV set, schemas and non-empty content."""
    directory = Path(directory)
    expected = expected_csv_filenames(period)
    actual = {path.name for path in directory.glob(f"{period}_*.csv")}
    missing = sorted(expected - actual)
    unexpected = sorted(actual - expected)
    if missing or unexpected:
        details = []
        if missing:
            details.append(f"faltan: {', '.join(missing)}")
        if unexpected:
            details.append(f"sobran: {', '.join(unexpected)}")
        raise ReportValidationError("Conjunto CSV inválido (" + "; ".join(details) + ").")

    validated = []
    for report_name, contract in CSV_CONTRACTS.items():
        path = directory / f"{period}_{report_name}.csv"
        if path.stat().st_size == 0:
            raise ReportValidationError(f"El archivo {path.name} está vacío.")

        dataframe = pd.read_csv(path, sep=contract.separator)
        if dataframe.empty:
            raise ReportValidationError(f"El archivo {path.name} no contiene registros.")
        if tuple(dataframe.columns) != contract.columns:
            raise ReportValidationError(f"El archivo {path.name} no cumple el esquema esperado.")
        validated.append(path)

    return validated


def validate_excel_outputs(directory: str | Path, period: str) -> list[Path]:
    """Validate the exact staged Excel set and basic workbook readability."""
    directory = Path(directory)
    expected = expected_excel_filenames(period)
    actual = {path.name for path in directory.glob(f"{period}_*.xlsx")}
    missing = sorted(expected - actual)
    unexpected = sorted(actual - expected)
    if missing or unexpected:
        details = []
        if missing:
            details.append(f"faltan: {', '.join(missing)}")
        if unexpected:
            details.append(f"sobran: {', '.join(unexpected)}")
        raise ReportValidationError("Conjunto Excel inválido (" + "; ".join(details) + ").")

    validated = []
    for filename in sorted(expected):
        path = directory / filename
        if path.stat().st_size == 0:
            raise ReportValidationError(f"El archivo {filename} está vacío.")

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if not workbook.sheetnames:
                raise ReportValidationError(f"El archivo {filename} no contiene hojas.")
            if not any(
                sheet.max_row > 0 and sheet.max_column > 0
                for sheet in workbook.worksheets
            ):
                raise ReportValidationError(f"El archivo {filename} no contiene datos.")
        finally:
            workbook.close()
        validated.append(path)

    return validated


def publish_staged_directories(
    staged_and_official: Iterable[tuple[str | Path, str | Path]],
) -> None:
    """Replace official directories and restore every prior output on failure."""
    pairs = [(Path(staged), Path(official)) for staged, official in staged_and_official]
    token = uuid.uuid4().hex
    backups: dict[Path, Path] = {}
    published: list[Path] = []

    try:
        for staged, official in pairs:
            if not staged.is_dir():
                raise ReportValidationError(f"No existe el directorio temporal {staged}.")
            official.parent.mkdir(parents=True, exist_ok=True)

            backup = official.with_name(f".{official.name}.backup-{token}")
            if official.exists():
                official.replace(backup)
                backups[official] = backup

            staged.replace(official)
            published.append(official)
    except Exception:
        for official in reversed(published):
            if official.exists():
                shutil.rmtree(official)
        for official, backup in backups.items():
            if backup.exists():
                backup.replace(official)
        raise
    else:
        for backup in backups.values():
            if backup.exists():
                shutil.rmtree(backup)


def extract_failed_reports(output: str) -> list[str]:
    """Extract report names from the functional CLI failure summary."""
    failed = []
    for line in output.splitlines():
        stripped = line.strip()
        if not stripped.startswith("❌") or "falló:" not in stripped:
            continue
        report_name = stripped.removeprefix("❌").split("falló:", 1)[0].strip()
        if report_name:
            failed.append(report_name)
    return failed


def generate_official_reports(
    project_root: str | Path,
    database_path: str | Path,
    period: int,
) -> dict:
    """Generate, validate and atomically publish CSV and Excel for one period."""
    validate_report_preflight(database_path, period)

    root = Path(project_root)
    period_text = str(period)
    logs: list[str] = []
    csv_script_path = root / "ending_files" / "generate_all_reports.py"
    excel_script_path = root / "excel_generators" / "generate_all_excel.py"
    official_csv_dir = root / "ending_files" / period_text
    official_excel_dir = root / "excel_final_files" / period_text

    try:
        with tempfile.TemporaryDirectory(
            prefix=".report-generation-",
            dir=root,
        ) as temporary_root:
            staging_root = Path(temporary_root)
            staging_csv_root = staging_root / "ending_files"
            staging_csv_dir = staging_csv_root / period_text
            staging_excel_dir = staging_root / "excel_final_files" / period_text

            logs.append("Iniciando generación de archivos CSV...")
            csv_result = subprocess.run(
                [
                    sys.executable,
                    str(csv_script_path),
                    period_text,
                    "--output_dir",
                    str(staging_csv_root),
                ],
                capture_output=True,
                text=True,
                cwd=str(root),
                timeout=300,
            )
            if csv_result.returncode != 0:
                failed = extract_failed_reports(csv_result.stdout)
                logging.error(
                    "Falló la generación CSV para %s. stdout=%s stderr=%s",
                    period_text,
                    csv_result.stdout,
                    csv_result.stderr,
                )
                if failed:
                    status = "partial" if len(failed) < len(CSV_CONTRACTS) else "failed"
                    message = (
                        f"No se generaron {len(failed)} archivos CSV: "
                        f"{', '.join(failed)}."
                    )
                else:
                    status = "failed"
                    message = "Los archivos CSV no se generaron."
                raise ReportGenerationFailure(
                    message,
                    status=status,
                    logs=logs,
                    failed_csv_reports=failed,
                )

            try:
                csv_files = validate_csv_outputs(staging_csv_dir, period_text)
            except ReportValidationError as error:
                logging.error(
                    "Falló la validación CSV para %s: %s", period_text, error
                )
                raise ReportGenerationFailure(
                    "Los archivos CSV no superaron los controles finales.",
                    logs=logs,
                ) from error

            logs.append(f"{len(csv_files)} archivos CSV generados y validados.")
            logs.append("Iniciando generación de archivos Excel...")
            excel_result = subprocess.run(
                [
                    sys.executable,
                    str(excel_script_path),
                    period_text,
                    "--csv-dir",
                    str(staging_csv_dir),
                    "--output-dir",
                    str(staging_excel_dir),
                ],
                capture_output=True,
                text=True,
                cwd=str(root),
                timeout=600,
            )
            if excel_result.returncode != 0:
                failed = extract_failed_reports(excel_result.stdout)
                logging.error(
                    "Falló la generación Excel para %s. stdout=%s stderr=%s",
                    period_text,
                    excel_result.stdout,
                    excel_result.stderr,
                )
                if failed:
                    status = (
                        "partial" if len(failed) < len(EXCEL_REPORT_NAMES) else "failed"
                    )
                    message = (
                        f"No se generaron {len(failed)} archivos Excel: "
                        f"{', '.join(failed)}."
                    )
                else:
                    status = "failed"
                    message = "Los archivos Excel no se generaron."
                raise ReportGenerationFailure(
                    message,
                    status=status,
                    logs=logs,
                    failed_excel_reports=failed,
                )

            try:
                excel_files = validate_excel_outputs(staging_excel_dir, period_text)
            except ReportValidationError as error:
                logging.error(
                    "Falló la validación Excel para %s: %s", period_text, error
                )
                raise ReportGenerationFailure(
                    "Los archivos Excel no superaron los controles finales.",
                    logs=logs,
                ) from error

            logs.append(f"{len(excel_files)} archivos Excel generados y validados.")
            publish_staged_directories(
                (
                    (staging_csv_dir, official_csv_dir),
                    (staging_excel_dir, official_excel_dir),
                )
            )
            logs.append("La salida oficial del período fue actualizada.")

        return {
            "status": "success",
            "logs": logs,
            "message": (
                f"Todos los reportes fueron generados y validados para {period_text}"
            ),
            "csv_directory": str(official_csv_dir),
            "excel_directory": str(official_excel_dir),
            "csv_count": len(csv_files),
            "excel_count": len(excel_files),
            "periodo": period_text,
        }
    except ReportGenerationFailure:
        raise
    except subprocess.TimeoutExpired as error:
        raise ReportGenerationFailure(
            "El proceso excedió el tiempo límite. La salida oficial no fue modificada.",
            logs=logs,
        ) from error
    except Exception as error:
        logging.exception("Error inesperado generando reportes para %s", period_text)
        raise ReportGenerationFailure(
            "La generación no pudo completarse. La salida oficial no fue modificada.",
            logs=logs,
        ) from error
