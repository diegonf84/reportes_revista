import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

def create_excel_cuadro_principal_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'cuadro_principal' con todas las hojas.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Convertir porcentajes de coma a punto
    pct_columns = ['pct_stros', 'pct_gastos', 'pct_result']
    for col in pct_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').astype(float)
    
    # Convertir siniestros y gastos (pueden tener coma decimal)
    numeric_columns = ['siniestros', 'gastos']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').astype(float)
    
    # Mapeos de nombres
    mapa_ramos_generales = {
        'Accidentes a Pasajeros': 'ACCIDENTES A PASAJEROS',
        'Aeronavegación': 'AERONAVEGACION',
        'Automotores': 'AUTOMOTORES',
        'Caución': 'CAUCION',
        'Combinado Familiar e Integral': 'COMBINADO FLIAR. E INTEGR.',
        'Créditos': 'CREDITO',
        'Incendio': 'INCENDIO',
        'Motovehículos': 'MOTOVEHICULOS',
        'Otros Riesgos de Daños Patrimoniales': 'OTROS RIESGOS PATRIMONIALES',
        'Responsabilidad Civil': 'RESPONSABILIDAD CIVIL',
        'Riesgos Agropecuarios y Forestales': 'RIESGOS AGROPEC. Y FOREST.',
        'Robo y Riesgos Similares': 'ROBO Y RIESGOS SIMILARES',
        'Técnico': 'TECNICO',
        'Transporte Público de Pasajeros': 'TRANSPORTE PUBL. DE PASAJ.',
        'Transportes - Cascos': 'TRANSPORTES - CASCOS',
        'Transportes - Mercaderías': 'TRANSPORTES - MERCADERIAS'
    }
    
    mapa_ramos_vida = {
        'Accidentes Personales': 'ACCIDENTES PERSONALES',
        'Salud': 'SALUD',
        'Sepelio': 'SEPELIO',
        'Vida': 'VIDA'
    }
    
    # Clasificar ramos por tipo
    todos_los_ramos = df['ramo_denominacion'].unique()
    
    ramos_generales = [ramo for ramo in todos_los_ramos if ramo in mapa_ramos_generales.keys()]
    ramos_vida = [ramo for ramo in todos_los_ramos if ramo in mapa_ramos_vida.keys()]
    ramos_art = [ramo for ramo in todos_los_ramos if ramo == 'Riesgos del Trabajo']
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear hojas
    if ramos_generales:
        generales_data = df[df['ramo_denominacion'].isin(ramos_generales)].copy()
        ws_generales = wb.create_sheet(title="Generales")
        crear_hoja_tipo(ws_generales, generales_data, ramos_generales, mapa_ramos_generales, tiene_resumen=True)
    
    if ramos_vida:
        vida_data = df[df['ramo_denominacion'].isin(ramos_vida)].copy()
        ws_vida = wb.create_sheet(title="Vida")
        crear_hoja_tipo(ws_vida, vida_data, ramos_vida, mapa_ramos_vida, tiene_resumen=True)
    
    if ramos_art:
        art_data = df[df['ramo_denominacion'].isin(ramos_art)].copy()
        ws_art = wb.create_sheet(title="ART")
        crear_hoja_tipo(ws_art, art_data, ramos_art, {'Riesgos del Trabajo': 'ART'}, tiene_resumen=False)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel completo generado con hojas: {[sheet.title for sheet in wb.worksheets]} en: {output_path}")

def crear_hoja_tipo(ws, data, ramos_disponibles, mapa_nombres, tiene_resumen=True):
    """Crea una hoja completa para un tipo (Generales, Vida, ART)"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Estilos
    title_font = Font(name="Arial", size=11, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # === TÍTULO INICIAL (siempre) ===
    ws.cell(row=current_row, column=1, value="RESULTADOS TECNICOS TOTALES POR RAMO").font = title_font
    current_row += 3  # Título + 2 líneas en blanco
    
    # === CUADRO RESUMEN (solo si tiene_resumen=True) ===
    if tiene_resumen:
        current_row = crear_cuadro_resumen(ws, data, ramos_disponibles, mapa_nombres, 
                                         current_row, title_font, header_font, normal_font, total_font, 
                                         center_alignment, center_vertical_alignment, thin_border)
        current_row += 3
    
    # === CUADROS INDIVIDUALES POR RAMO ===
    for ramo in ramos_disponibles:
        datos_ramo = data[data['ramo_denominacion'] == ramo].copy().sort_values('primas_emitidas', ascending=False)
        nombre_ramo_excel = mapa_nombres[ramo]
        
        current_row = crear_cuadro_ramo(ws, datos_ramo, nombre_ramo_excel, current_row,
                                      title_font, header_font, normal_font, total_font,
                                      center_alignment, center_vertical_alignment, thin_border)
        current_row += 3
    
    # Ajustar anchos de columnas
    column_widths = [33, 15, 15, 12, 18, 12, 8]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def crear_cuadro_resumen(ws, data, ramos_disponibles, mapa_nombres, start_row, 
                        title_font, header_font, normal_font, total_font, center_alignment, 
                        center_vertical_alignment, thin_border):
    """Crea el cuadro resumen por ramos"""
    
    current_row = start_row
    
    # Headers en 2 filas (sin título, ya se agregó en crear_hoja_tipo)
    current_row = crear_headers_dobles(ws, current_row, header_font, center_alignment,
                                     center_vertical_alignment, thin_border, es_cuadro_resumen=True)
    
    # Datos por ramo
    for ramo in ramos_disponibles:
        datos_ramo = data[data['ramo_denominacion'] == ramo]
        nombre_excel = mapa_nombres[ramo]
        
        # Calcular totales del ramo
        totales = calcular_totales_ramo(datos_ramo)
        
        # Escribir fila del ramo
        cell_ramo = ws.cell(row=current_row, column=1, value=nombre_excel)
        cell_ramo.font = normal_font
        cell_ramo.border = thin_border
        
        valores = [
            totales['primas_emitidas'],
            totales['primas_devengadas'],
            round(totales['siniestros'] / totales['primas_devengadas'] * 100, 1),
            round(totales['gastos'] / totales['primas_devengadas'] * 100, 1),
            totales['resultado'],
            round(totales['resultado'] / totales['primas_devengadas'] * 100, 1)
        ]
        
        for col_idx, valor in enumerate(valores, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=valor)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear números
            if col_idx in [2, 3, 6]:  # Primas emitidas, devengadas, resultado
                cell.number_format = '#,##0,'
            elif col_idx in [4, 5]:  # Porcentajes
                cell.number_format = '0.0'
            elif col_idx == 7:  # Porcentaje resultado
                cell.number_format = '0.0'
            
            cell.alignment = center_alignment
        
        current_row += 1
    
    # Calcular total general usando la función específica
    total_general = calcular_total_general(data, ramos_disponibles)
    
    # Fila TOTAL DEL MERCADO
    cell_total = ws.cell(row=current_row, column=1, value="TOTAL DEL MERCADO")
    cell_total.font = total_font
    cell_total.border = thin_border
    
    valores_total = [
        total_general['primas_emitidas'],
        total_general['primas_devengadas'],
        round(total_general['siniestros'] / total_general['primas_devengadas'] * 100, 1),
        round(total_general['gastos'] / total_general['primas_devengadas'] * 100, 1),
        total_general['resultado'],
        round(total_general['resultado'] / total_general['primas_devengadas'] * 100, 1)
    ]
    
    for col_idx, valor in enumerate(valores_total, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        
        if col_idx in [2, 3, 6]:
            cell.number_format = '#,##0,'
        elif col_idx in [4, 5, 7]:
            cell.number_format = '0.0'
        
        cell.alignment = center_alignment
    
    return current_row + 1

def crear_cuadro_ramo(ws, datos_ramo, nombre_ramo, start_row, title_font, 
                     header_font, normal_font, total_font, center_alignment, 
                     center_vertical_alignment, thin_border):
    """Crea cuadro individual para un ramo"""
    
    current_row = start_row
    
    # Título del ramo
    ws.cell(row=current_row, column=1, value=nombre_ramo).font = title_font
    current_row += 2
    
    # Headers en 2 filas
    current_row = crear_headers_dobles(ws, current_row, header_font, center_alignment,
                                     center_vertical_alignment, thin_border, es_cuadro_resumen=False)
    
    # Ordenar datos por primas emitidas de mayor a menor
    datos_ordenados = datos_ramo.sort_values('primas_emitidas', ascending=False)
    
    # Datos de entidades
    for _, row_data in datos_ordenados.iterrows():
        cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
        cell_entidad.font = normal_font
        cell_entidad.border = thin_border
        
        valores = [
            row_data['primas_emitidas'],
            row_data['primas_devengadas'],
            row_data['pct_stros'],
            row_data['pct_gastos'],
            row_data['resultado'],
            row_data['pct_result']
        ]
        
        for col_idx, valor in enumerate(valores, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=valor)
            cell.font = normal_font
            cell.border = thin_border
            
            if col_idx in [2, 3, 6]:
                cell.number_format = '#,##0,'
            elif col_idx in [4, 5, 7]:
                cell.number_format = '0.0'
            
            cell.alignment = center_alignment
        
        current_row += 1
    
    # Total del ramo
    totales = calcular_totales_ramo(datos_ramo)
    
    cell_total = ws.cell(row=current_row, column=1, value="TOTAL")
    cell_total.font = total_font
    cell_total.border = thin_border
    
    valores_total = [
        totales['primas_emitidas'],
        totales['primas_devengadas'],
        round(totales['siniestros'] / totales['primas_devengadas'] * 100, 1),
        round(totales['gastos'] / totales['primas_devengadas'] * 100, 1),
        totales['resultado'],
        round(totales['resultado'] / totales['primas_devengadas'] * 100, 1)
    ]
    
    for col_idx, valor in enumerate(valores_total, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        
        if col_idx in [2, 3, 6]:
            cell.number_format = '#,##0,'
        elif col_idx in [4, 5, 7]:
            cell.number_format = '0.0'
        
        cell.alignment = center_alignment
    
    return current_row + 1

def crear_headers_dobles(ws, start_row, header_font, center_alignment, 
                        center_vertical_alignment, thin_border, es_cuadro_resumen=True):
    """Crea headers en 2 filas con combinaciones"""
    
    # Título de primera columna según el tipo de cuadro
    primera_columna = "RAMO" if es_cuadro_resumen else "ENTIDAD"
    
    # Fila 1 - Headers principales
    headers_fila1 = [primera_columna, "Primas emitidas", "Primas devengadas", "Siniestr. (%)", 
                     "Gs. prod. y explot. (%)", "Resultados técnicos"]
    
    for col_idx, header in enumerate(headers_fila1, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = center_vertical_alignment  # Primera columna con centrado vertical
        elif col_idx > 1:
            cell.alignment = center_vertical_alignment
    
    # Combinar celda F-G para "Resultados técnicos"
    ws.merge_cells(f'F{start_row}:G{start_row}')
    
    # Fila 2 - Subheaders para Resultados técnicos
    ws.cell(row=start_row + 1, column=6, value="$").font = header_font
    ws.cell(row=start_row + 1, column=6).border = thin_border
    ws.cell(row=start_row + 1, column=6).alignment = center_vertical_alignment
    
    ws.cell(row=start_row + 1, column=7, value="%").font = header_font
    ws.cell(row=start_row + 1, column=7).border = thin_border
    ws.cell(row=start_row + 1, column=7).alignment = center_vertical_alignment
    
    # Extender celdas de la fila 1 a la fila 2 (A-E) y aplicar bordes
    for col_idx in range(1, 6):
        # Aplicar borde a la celda de la segunda fila antes de combinar
        cell_fila2 = ws.cell(row=start_row + 1, column=col_idx)
        cell_fila2.border = thin_border
        
        # Combinar las celdas
        ws.merge_cells(f'{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{start_row + 1}')
    
    # Ajustar altura de ambas filas
    ws.row_dimensions[start_row].height = 21.5
    ws.row_dimensions[start_row + 1].height = 21.5
    
    return start_row + 2

def calcular_totales_ramo(datos_ramo):
    """Calcula totales para un ramo"""
    return {
        'primas_emitidas': datos_ramo['primas_emitidas'].sum(),
        'primas_devengadas': datos_ramo['primas_devengadas'].sum(),
        'siniestros': datos_ramo['siniestros'].sum(),
        'gastos': datos_ramo['gastos'].sum(),
        'resultado': datos_ramo['resultado'].sum()
    }

def calcular_total_general(data, ramos_disponibles):
    """Calcula totales generales para todos los ramos"""
    total_general = {
        'primas_emitidas': 0, 'primas_devengadas': 0, 'siniestros': 0,
        'gastos': 0, 'resultado': 0
    }
    
    for ramo in ramos_disponibles:
        datos_ramo = data[data['ramo_denominacion'] == ramo]
        totales_ramo = calcular_totales_ramo(datos_ramo)
        
        for key in total_general:
            total_general[key] += totales_ramo[key]
    
    return total_general

def generate_cuadro_principal_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel completo de cuadro principal.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_cuadro_principal.csv")
    output_path = os.path.join(period_dir, f"{period}_cuadro_principal.xlsx")
    
    create_excel_cuadro_principal_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel completo para cuadro principal')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_cuadro_principal_excel(args.period)
        print(f"✅ Excel completo generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_cuadro_principal.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")