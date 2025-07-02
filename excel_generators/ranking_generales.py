import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

def create_excel_ranking_produccion(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel de ranking de producción con 3 hojas.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Filtrar tipos
    tipos_incluir = ['ART', 'Generales', 'M.T.P.P.', 'Retiro','Vida']
    df_filtrado = df[df['tipo_cia'].isin(tipos_incluir)].copy()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === HOJA 1: RANKING GENERALES ===
    ws1 = wb.create_sheet(title="Ranking Generales")
    crear_hoja_ranking_generales(ws1, df_filtrado)
    
    # === HOJA 2: VARIOS ===
    ws2 = wb.create_sheet(title="Varios")
    crear_hoja_varios(ws2, df_filtrado, tipos_incluir)
    
    # === HOJA 3: RANKING ===
    ws3 = wb.create_sheet(title="Ranking")
    crear_hoja_ranking_total(ws3, df_filtrado)
    
    wb.save(output_path)
    logging.info(f"Excel de ranking de producción generado en: {output_path}")

def crear_hoja_ranking_generales(ws, df):
    """Crea hoja 'Ranking Generales'"""
    
    ws.sheet_view.showGridLines = False
    
    # Estilos
    title_font = Font(name="Arial", size=12, bold=True)  # Tamaño 12
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título en A1
    ws.cell(row=1, column=1, value="RANKING DE PRODUCCION DE COMPAÑIAS DE SEGUROS PATRIMONIALES Y MIXTAS").font = title_font
    
    # Datos solo de Generales, ordenados de menor a mayor
    generales_data = df[df['tipo_cia'] == 'Generales'].copy()
    generales_ordenados = generales_data.sort_values('primas_emitidas', ascending=False)
    
    current_row = 3
    
    # Headers
    headers = ["ENTIDAD", "Primas emitidas"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx > 1:
            cell.alignment = center_alignment
    current_row += 1
    
    # Datos
    for _, row_data in generales_ordenados.iterrows():
        # ENTIDAD
        cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
        cell_entidad.font = normal_font
        cell_entidad.border = thin_border
        
        # PRIMAS
        cell_primas = ws.cell(row=current_row, column=2, value=row_data['primas_emitidas'])
        cell_primas.font = normal_font
        cell_primas.border = thin_border
        cell_primas.number_format = '#,##0,'
        cell_primas.alignment = center_alignment
        
        current_row += 1
    
    # 2 líneas vacías
    current_row += 2
    
    # Total del mercado
    total_primas = generales_data['primas_emitidas'].sum()
    
    cell_total_label = ws.cell(row=current_row, column=1, value="TOTAL DEL MERCADO")
    cell_total_label.font = total_font
    cell_total_label.border = thin_border
    
    cell_total_value = ws.cell(row=current_row, column=2, value=total_primas)
    cell_total_value.font = total_font
    cell_total_value.border = thin_border
    cell_total_value.number_format = '#,##0,'
    cell_total_value.alignment = center_alignment
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 39
    ws.column_dimensions['B'].width = 16

def crear_hoja_varios(ws, df, tipos_incluir):
    """Crea hoja 'Varios' con cuadros por tipo y resumen"""
    
    ws.sheet_view.showGridLines = False
    
    # Estilos
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 2  # Empezar en fila 2 (sin título)
    
    # Cuadros por tipo
    for tipo in tipos_incluir:
        tipo_data = df[df['tipo_cia'] == tipo].copy()
        if tipo_data.empty:
            continue
            
        nombre_tipo = "MTTP" if tipo == "M.T.P.P." else tipo
        
        # Título del tipo
        ws.cell(row=current_row, column=1, value=nombre_tipo).font = title_font
        current_row += 2  # Línea en blanco
        
        # Headers
        headers = ["ENTIDAD", "Primas emitidas", "% del rubro", "% acumulado"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_alignment
        current_row += 1
        
        # Ordenar de mayor a menor
        tipo_ordenado = tipo_data.sort_values('primas_emitidas', ascending=False)
        total_tipo = tipo_data['primas_emitidas'].sum()
        porcentaje_acumulado = 0
        
        # Datos
        for _, row_data in tipo_ordenado.iterrows():
            porcentaje_rubro = (row_data['primas_emitidas'] / total_tipo) * 100
            porcentaje_acumulado += porcentaje_rubro
            
            # ENTIDAD
            cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # PRIMAS
            cell_primas = ws.cell(row=current_row, column=2, value=row_data['primas_emitidas'])
            cell_primas.font = normal_font
            cell_primas.border = thin_border
            cell_primas.number_format = '#,##0,'
            cell_primas.alignment = center_alignment
            
            # % DEL RUBRO
            cell_pct = ws.cell(row=current_row, column=3, value=porcentaje_rubro)
            cell_pct.font = normal_font
            cell_pct.border = thin_border
            cell_pct.number_format = '0.00'
            cell_pct.alignment = center_alignment
            
            # % ACUMULADO
            cell_acum = ws.cell(row=current_row, column=4, value=porcentaje_acumulado)
            cell_acum.font = normal_font
            cell_acum.border = thin_border
            cell_acum.number_format = '0.00'
            cell_acum.alignment = center_alignment
            
            current_row += 1
        
        # Fila Total del tipo
        cell_total_label = ws.cell(row=current_row, column=1, value="Total")
        cell_total_label.font = total_font
        cell_total_label.border = thin_border
        
        cell_total_primas = ws.cell(row=current_row, column=2, value=total_tipo)
        cell_total_primas.font = total_font
        cell_total_primas.border = thin_border
        cell_total_primas.number_format = '#,##0,'
        cell_total_primas.alignment = center_alignment
        
        # No agregar valores en columnas C y D para el total
        
        current_row += 1
        
        current_row += 2  # Espacio entre tipos
    
    # === CUADRO RESUMEN EN COLUMNA H ===
    crear_cuadro_resumen(ws, df, tipos_incluir)
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 11.5
    ws.column_dimensions['D'].width = 11.5
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 8

def crear_cuadro_resumen(ws, df, tipos_incluir):
    """Crea cuadro resumen en columna H fila 3"""
    
    # Estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 3
    
    # Headers del resumen
    headers_resumen = ["RUBRO", "Primas emitidas", "%"]
    for col_idx, header in enumerate(headers_resumen, 8):  # Columna H=8
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx > 8:
            cell.alignment = center_alignment
    current_row += 1
    
    # Calcular total general
    total_general = df['primas_emitidas'].sum()
    
    # Datos por tipo
    for tipo in tipos_incluir:
        tipo_data = df[df['tipo_cia'] == tipo].copy()
        if tipo_data.empty:
            continue
            
        total_tipo = tipo_data['primas_emitidas'].sum()
        porcentaje_tipo = (total_tipo / total_general) * 100
        nombre_tipo = "MTTP" if tipo == "M.T.P.P." else tipo
        
        # RUBRO
        cell_rubro = ws.cell(row=current_row, column=8, value=nombre_tipo)
        cell_rubro.font = normal_font
        cell_rubro.border = thin_border
        
        # PRIMAS
        cell_primas = ws.cell(row=current_row, column=9, value=total_tipo)
        cell_primas.font = normal_font
        cell_primas.border = thin_border
        cell_primas.number_format = '#,##0,'
        cell_primas.alignment = center_alignment
        
        # %
        cell_pct = ws.cell(row=current_row, column=10, value=porcentaje_tipo)
        cell_pct.font = normal_font
        cell_pct.border = thin_border
        cell_pct.number_format = '0.00'
        cell_pct.alignment = center_alignment
        
        current_row += 1
    
    # Total
    cell_total_label = ws.cell(row=current_row, column=8, value="Total")
    cell_total_label.font = total_font
    cell_total_label.border = thin_border
    
    cell_total_primas = ws.cell(row=current_row, column=9, value=total_general)
    cell_total_primas.font = total_font
    cell_total_primas.border = thin_border
    cell_total_primas.number_format = '#,##0,'
    cell_total_primas.alignment = center_alignment
    
    cell_total_pct = ws.cell(row=current_row, column=10, value=100.0)
    cell_total_pct.font = total_font
    cell_total_pct.border = thin_border
    cell_total_pct.number_format = '0.00'
    cell_total_pct.alignment = center_alignment

def crear_hoja_ranking_total(ws, df):
    """Crea hoja 'Ranking' con todas las entidades"""
    
    ws.sheet_view.showGridLines = False
    
    # Estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 3  # Empezar en fila 3 (2 líneas en blanco)
    
    # Headers
    headers = ["ENTIDAD", "Primas emitidas"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx > 1:
            cell.alignment = center_alignment
    current_row += 1
    
    # Todas las entidades ordenadas de mayor a menor
    todas_ordenadas = df.sort_values('primas_emitidas', ascending=False)
    
    # Datos
    for _, row_data in todas_ordenadas.iterrows():
        # ENTIDAD
        cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
        cell_entidad.font = normal_font
        cell_entidad.border = thin_border
        
        # PRIMAS
        cell_primas = ws.cell(row=current_row, column=2, value=row_data['primas_emitidas'])
        cell_primas.font = normal_font
        cell_primas.border = thin_border
        cell_primas.number_format = '#,##0,'
        cell_primas.alignment = center_alignment
        
        current_row += 1
    
    # 2 líneas vacías
    current_row += 2
    
    # Total mercado
    total_mercado = df['primas_emitidas'].sum()
    
    cell_total_label = ws.cell(row=current_row, column=1, value="TOTAL MERCADO")
    cell_total_label.font = total_font
    cell_total_label.border = thin_border
    
    cell_total_value = ws.cell(row=current_row, column=2, value=total_mercado)
    cell_total_value.font = total_font
    cell_total_value.border = thin_border
    cell_total_value.number_format = '#,##0,'
    cell_total_value.alignment = center_alignment
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 39
    ws.column_dimensions['B'].width = 16

def generate_ranking_produccion_excel(period: str) -> str:
    """
    Genera Excel de ranking de producción.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_ranking_comparativo.csv")
    output_path = os.path.join(period_dir, f"{period}_ranking_produccion.xlsx")
    
    create_excel_ranking_produccion(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel de ranking de producción')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_ranking_produccion_excel(args.period)
        print(f"✅ Excel de ranking generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_ranking_comparativo.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")