import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import os
import logging
import argparse

def create_excel_detalle_inmuebles(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel formateado para el reporte 'detalle_inmuebles'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path)
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Detalle Inmuebles {period}"
    
    # Quitar las líneas de cuadrícula de la hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=10, bold=True, underline="single") # Fuente para títulos de sección (ART, etc.)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    # Definir alineación
    center_alignment = Alignment(horizontal='center')
    
    # Definir bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers de columnas, comenzando en columna B
    headers = [
        "ENTIDAD", "Inmuebles Inversión", "Inmuebles Uso Propio", "Total Inmuebles"
    ]
    
    # Mapeo de columnas del CSV a headers del Excel
    column_mapping = {
        'nombre_corto': 'ENTIDAD',
        'inmuebles_inversion': 'Inmuebles Inversión',
        'inmuebles_uso_propio': 'Inmuebles Uso Propio',
        'inmuebles_total': 'Total Inmuebles'
    }
    
    # Columnas numéricas
    numerical_columns = [col for col in column_mapping if col != 'nombre_corto']
    
    current_row = 1
    
    # Procesar cada tipo de compañía
    for tipo_cia in df['tipo_cia'].unique():
        
        # Filtrar datos para este tipo
        data_tipo = df[df['tipo_cia'] == tipo_cia].copy()
        
        # Título del tipo (ej: "ART", "Generales") en columna A
        cell_tipo = ws.cell(row=current_row, column=1, value=tipo_cia)
        cell_tipo.font = title_font # Aplicar fuente con subrayado
        current_row += 1
        
        # Headers de columnas, comenzando en columna B
        for col_idx, header in enumerate(headers, 2):  # Empezar en columna 2 (B)
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            # Aplicar alineación centralizada de la columna C en adelante
            if col_idx > 2:
                cell.alignment = center_alignment
        current_row += 1
        
        # Datos de las compañías, ordenados por inmuebles_total descendente
        data_tipo_ordenado = data_tipo.sort_values('inmuebles_total', ascending=False)
        for _, row_data in data_tipo_ordenado.iterrows():
            # Nombre de la entidad en columna B
            cell_entidad = ws.cell(row=current_row, column=2, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # Valores numéricos comenzando en columna C
            for col_idx, csv_col in enumerate(numerical_columns, 3):  # Empezar en columna 3 (C)
                value = row_data[csv_col]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment # Aplicar alineación centralizada
            current_row += 1
        
        # Fila de totales
        cell_total = ws.cell(row=current_row, column=2, value="Total")
        cell_total.font = total_font
        cell_total.border = thin_border
        
        # Calcular y escribir totales, comenzando en columna C
        for col_idx, csv_col in enumerate(numerical_columns, 3):  # Empezar en columna 3 (C)
            total_value = data_tipo[csv_col].sum()
            cell = ws.cell(row=current_row, column=col_idx, value=total_value)
            cell.font = total_font
            cell.border = thin_border
            cell.number_format = '#,##0,'
            cell.alignment = center_alignment # Aplicar alineación centralizada
        
        current_row += 2  # Espacio entre tipos
    
    # Ajustar ancho de columnas
    column_widths = [15, 33.5, 18, 18, 18]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel detalle inmuebles generado en: {output_path}")

def generate_detalle_inmuebles_excel(period: str, csv_dir: str = None) -> str:
    """
    Función de conveniencia para generar el Excel del detalle de inmuebles.
    
    Args:
        period: Período (ej: "202501")
        csv_dir: Directorio donde está el CSV (default: ending_files/{period}/)
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    if csv_dir is None:
        csv_dir = os.path.join(base_dir, "ending_files", period)
    
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_detalle_inmuebles.csv")
    output_path = os.path.join(period_dir, f"{period}_detalle_inmuebles.xlsx")
    
    create_excel_detalle_inmuebles(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Genera Excel formateado para detalle de inmuebles')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    parser.add_argument('--csv_dir', default=None, help='Directorio donde está el CSV')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    try:
        excel_path = generate_detalle_inmuebles_excel(
            period=args.period,
            csv_dir=args.csv_dir
        )
        print(f"✅ Excel generado exitosamente: {excel_path}")
    except FileNotFoundError as e:
        print(f"❌ Error: No se encontró el archivo CSV esperado")
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        expected_csv_path = f"{base_dir}/ending_files/{args.period}" if args.csv_dir is None else args.csv_dir
        print(f"   Buscando CSV en: {expected_csv_path}/{args.period}_detalle_inmuebles.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")