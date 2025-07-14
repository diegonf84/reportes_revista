import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

# Diccionario de mapeo de subramos (fácil de modificar en el futuro)
SUBRAMOS_INCLUIDOS = {
    "Comb. Fam. e Int. - Comb. Fam.": "Combinado Familiar",
    "Comb. Fam. e Int. - Int. de Com.": "Integral de Comercio",
    "Comb. Fam. e Int. - Otros": "Otros",
    "Automotores - Cascos y Otras Cob.": "Cascos y Otras Coberturas",
    "Automotores - RC Exclusivo": "RC Exclusivo",
    "RC - Mala Práctica Médica": "Mala Praxis Médica",
    "RC - Profesional Otras Profesiones": "Otras Profesiones",
    "RC - Accidentes de Trabajo": "Accidentes de Trabajo",
    "RC - Otros": "Otros",
    "Ot. Rs. de Ds Patrim. - Cristales": "Cristales",
    "Ot. Rs. de Ds Patrim. - R.Varios": "Riesgos Varios",
    "Ot. Rs. de Ds Patrim. - Otros": "Otros",
    "Motos - Cascos y Otras Cob.": "Cascos y Otras Coberturas",
    "Motos - RC Exclusivo": "RC Exclusivo",
    "Acc. Personales - Individual": "Individual",
    "Acc. Personales - Colectivo": "Colectivo",
    "Salud - Individual": "Individual",
    "Salud - Colectivo": "Colectivo",
    "Vida - Individual": "Individual",
    "Vida - Colectivo": "Colectivo",
    "Vida - Obligatorios": "Obligatorios",
    "Vida - Saldo Deudor": "Saldo Deudor",
    "Sepelio - Individual": "Individual",
    "Sepelio - Colectivo": "Colectivo",
    "Retiro - Individual": "Individual",
    "Retiro - Colectivo": "Colectivo"
}

# Mapeo de nombres de hojas (para casos especiales)
NOMBRES_HOJAS = {
    "Ot. Rs. de Ds Patrim.": "Otros Riesgos",
    "RC": "Responsabilidad Civil",
    "Acc. Personales": "Accidentes Personales",
    "Comb. Fam. e Int.": "Combinado Familiar",
    "Motos": "Motovehículos"
}

def create_excel_apertura_subramo_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'apertura_por_subramo'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Filtrar solo los subramos incluidos
    df_filtrado = df[df['subramo_denominacion'].isin(SUBRAMOS_INCLUIDOS.keys())].copy()
    
    # Agregar columna de ramo (extraer texto antes del guión)
    df_filtrado['ramo'] = df_filtrado['subramo_denominacion'].apply(extraer_ramo)
    
    # Aplicar mapeo de nombres de subramos
    df_filtrado['subramo_display'] = df_filtrado['subramo_denominacion'].map(SUBRAMOS_INCLUIDOS)
    
    # Obtener ramos únicos
    ramos = df_filtrado['ramo'].unique()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear una hoja para cada ramo
    for ramo in sorted(ramos):
        # Aplicar mapeo de nombres de hoja si existe
        nombre_hoja = NOMBRES_HOJAS.get(ramo, ramo)
        
        # Truncar nombre si es muy largo para Excel
        if len(nombre_hoja) > 31:
            nombre_hoja = nombre_hoja[:31]
        
        ramo_data = df_filtrado[df_filtrado['ramo'] == ramo].copy()
        ws = wb.create_sheet(title=nombre_hoja)
        crear_hoja_ramo(ws, ramo_data, ramo)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel generado con {len(ramos)} hojas en: {output_path}")

def extraer_ramo(subramo_denominacion: str) -> str:
    """Extrae el nombre del ramo (texto antes del primer guión)"""
    if " - " in subramo_denominacion:
        return subramo_denominacion.split(" - ")[0].strip()
    return subramo_denominacion

def crear_hoja_ramo(ws, ramo_data, ramo_nombre):
    """Crea una hoja individual para un ramo con todos sus subramos"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Obtener subramos únicos de este ramo
    subramos = ramo_data['subramo_display'].unique()
    
    # Crear cuadro para cada subramo
    for subramo in sorted(subramos):
        subramo_data = ramo_data[ramo_data['subramo_display'] == subramo].copy()
        
        # Título del subramo
        ws.cell(row=current_row, column=1, value=subramo).font = title_font
        current_row += 2  # Línea en blanco
        
        # Headers
        headers = ["ENTIDAD", "Producción", "% del total", "% / total acum."]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_vertical_alignment
            else:
                cell.alignment = center_vertical_alignment
        
        # Ajustar altura de headers
        ws.row_dimensions[current_row].height = 39
        current_row += 1
        
        # Ordenar datos por primas descendente
        subramo_ordenado = subramo_data.sort_values('primas', ascending=False)
        total_subramo = subramo_data['primas'].sum()
        porcentaje_acumulado = 0
        
        # Datos de las entidades
        for _, row_data in subramo_ordenado.iterrows():
            porcentaje_del_total = (row_data['primas'] / total_subramo) * 100
            porcentaje_acumulado += porcentaje_del_total
            
            # ENTIDAD
            cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # PRODUCCIÓN
            cell_primas = ws.cell(row=current_row, column=2, value=row_data['primas'])
            cell_primas.font = normal_font
            cell_primas.border = thin_border
            cell_primas.number_format = '#,##0,'
            cell_primas.alignment = center_alignment
            
            # % DEL TOTAL
            cell_pct = ws.cell(row=current_row, column=3, value=porcentaje_del_total)
            cell_pct.font = normal_font
            cell_pct.border = thin_border
            cell_pct.number_format = '0.00'
            cell_pct.alignment = center_alignment
            
            # % ACUMULADO
            cell_acum = ws.cell(row=current_row, column=4, value=porcentaje_acumulado)
            cell_acum.font = normal_font
            cell_acum.border = thin_border
            cell_acum.number_format = '0.00'
            cell_acum.alignment = center_alignment
            
            current_row += 1
        
        # Fila Total del subramo
        cell_total_label = ws.cell(row=current_row, column=1, value="Total")
        cell_total_label.font = total_font
        cell_total_label.border = thin_border
        
        cell_total_primas = ws.cell(row=current_row, column=2, value=total_subramo)
        cell_total_primas.font = total_font
        cell_total_primas.border = thin_border
        cell_total_primas.number_format = '#,##0,'
        cell_total_primas.alignment = center_alignment
        
        # No agregar valores en columnas 3 y 4 para el total (como en ranking_generales)
        # Y quitar bordes de estas columnas en la fila Total
        cell_col3 = ws.cell(row=current_row, column=3, value="")
        # No aplicar borde a columna 3 en fila Total
        
        cell_col4 = ws.cell(row=current_row, column=4, value="")
        # No aplicar borde a columna 4 en fila Total
        
        current_row += 3  # Espacio entre subramos
    
    # Ajustar anchos de columnas
    column_widths = [35, 17, 11.5, 11.5]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def generate_apertura_subramo_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel de apertura por subramo.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_apertura_por_subramo.csv")
    output_path = os.path.join(period_dir, f"{period}_apertura_por_subramo.xlsx")
    
    create_excel_apertura_subramo_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel para apertura por subramo')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_apertura_subramo_excel(args.period)
        print(f"✅ Excel generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_apertura_por_subramo.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")