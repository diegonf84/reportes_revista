import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

SUBRAMOS_INCLUIDOS = {
    "Comb. Fam. e Int. - Comb. Fam.": "Combinado Familiar",
    "Comb. Fam. e Int. - Int. de Com.": "Integral de Comercio",
    "Comb. Fam. e Int. - Otros": "Otros",
    "Automotores - Cascos y Otras Cob.": "Cascos y Otras Coberturas",
    "Automotores - RC Exclusivo": "RC Exclusivo",
    "RC - Mala Práctica Médica": "Mala Praxis Médica",
    "RC - Profesional Otras Profesiones": "Otras Profesiones",
    "RC - Accidentes de Trabajo": "Accidentes de Trabajo",
    "RC - Otros": "Otros",
    "Ot. Rs. de Ds Patrim. - Cristales": "Cristales",
    "Ot. Rs. de Ds Patrim. - R.Varios": "Riesgos Varios",
    "Ot. Rs. de Ds Patrim. - Otros": "Otros",
    "Motos - Cascos y Otras Cob.": "Cascos y Otras Coberturas",
    "Motos - RC Exclusivo": "RC Exclusivo",
    "Acc. Personales - Individual": "Individual",
    "Acc. Personales - Colectivo": "Colectivo",
    "Salud - Individual": "Individual",
    "Salud - Colectivo": "Colectivo",
    "Vida - Individual": "Individual",
    "Vida - Colectivo": "Colectivo",
    "Vida - Obligatorios": "Obligatorios",
    "Vida - Saldo Deudor": "Saldo Deudor",
    "Sepelio - Individual": "Individual",
    "Sepelio - Colectivo": "Colectivo",
    "Retiro - Individual": "Individual",
    "Retiro - Colectivo": "Colectivo"
}

NOMBRES_HOJAS = {
    "Ot. Rs. de Ds Patrim.": "Otros Riesgos",
    "RC": "Responsabilidad Civil",
    "Acc. Personales": "Accidentes Personales",
    "Comb. Fam. e Int.": "Combinado Familiar",
    "Motos": "Motovehículos"
}

def create_excel_apertura_subramo_comparativo(csv_path: str, output_path: str, period: str) -> None:
    df = pd.read_csv(csv_path, sep=';')

    df['variacion'] = df['variacion'].astype(str).str.replace(',', '.').astype(float)

    df_filtrado = df[df['subramo_denominacion'].isin(SUBRAMOS_INCLUIDOS.keys())].copy()
    df_filtrado['ramo'] = df_filtrado['subramo_denominacion'].apply(extraer_ramo)
    df_filtrado['subramo_display'] = df_filtrado['subramo_denominacion'].map(SUBRAMOS_INCLUIDOS)

    ramos = df_filtrado['ramo'].unique()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ramo in sorted(ramos):
        nombre_hoja = NOMBRES_HOJAS.get(ramo, ramo)
        if len(nombre_hoja) > 31:
            nombre_hoja = nombre_hoja[:31]

        ramo_data = df_filtrado[df_filtrado['ramo'] == ramo].copy()
        ws = wb.create_sheet(title=nombre_hoja)
        crear_hoja_ramo(ws, ramo_data, ramo)

    ws_detalle = wb.create_sheet(title="base_detalle")
    crear_hoja_detalle(ws_detalle, df_filtrado)

    wb.save(output_path)
    logging.info(f"Excel generado con {len(ramos)} hojas + base_detalle en: {output_path}")


def extraer_ramo(subramo_denominacion: str) -> str:
    if " - " in subramo_denominacion:
        return subramo_denominacion.split(" - ")[0].strip()
    return subramo_denominacion


def _estilos():
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return (
        Font(name="Arial", size=10, bold=True, underline="single"),
        Font(name="Arial", size=10, bold=True),
        Font(name="Arial", size=10),
        Font(name="Arial", size=10, bold=True),
        Alignment(horizontal='center'),
        Alignment(horizontal='center', vertical='center', wrap_text=True),
        thin_border,
    )


def crear_hoja_ramo(ws, ramo_data, ramo_nombre):
    ws.sheet_view.showGridLines = False

    title_font, header_font, normal_font, total_font, center_alignment, center_vertical_alignment, thin_border = _estilos()

    current_row = 1
    subramos = ramo_data['subramo_display'].unique()

    for subramo in sorted(subramos):
        subramo_data = ramo_data[ramo_data['subramo_display'] == subramo].copy()

        ws.cell(row=current_row, column=1, value=subramo).font = title_font
        current_row += 2

        headers = ["ENTIDAD", "Producción", "Var %", "% del total", "% / total acum."]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_vertical_alignment
            if col_idx != 3:  # Var % sin borde (igual que ranking_comparativo_por_ramo)
                cell.border = thin_border

        ws.row_dimensions[current_row].height = 39
        current_row += 1

        subramo_ordenado = subramo_data.sort_values('primas', ascending=False)
        total_subramo = subramo_data['primas'].sum()
        porcentaje_acumulado = 0

        for _, row_data in subramo_ordenado.iterrows():
            porcentaje_del_total = (row_data['primas'] / total_subramo) * 100 if total_subramo else 0
            porcentaje_acumulado += porcentaje_del_total

            cell = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
            cell.font = normal_font
            cell.border = thin_border

            cell = ws.cell(row=current_row, column=2, value=row_data['primas'])
            cell.font = normal_font
            cell.border = thin_border
            cell.number_format = '#,##0,'
            cell.alignment = center_alignment

            cell = ws.cell(row=current_row, column=3, value=row_data['variacion'])
            cell.font = normal_font
            cell.number_format = '0.00'
            cell.alignment = center_alignment

            cell = ws.cell(row=current_row, column=4, value=porcentaje_del_total)
            cell.font = normal_font
            cell.border = thin_border
            cell.number_format = '0.00'
            cell.alignment = center_alignment

            cell = ws.cell(row=current_row, column=5, value=porcentaje_acumulado)
            cell.font = normal_font
            cell.border = thin_border
            cell.number_format = '0.00'
            cell.alignment = center_alignment

            current_row += 1

        # Fila Total
        totales = calcular_totales_subramo(subramo_data)

        cell = ws.cell(row=current_row, column=1, value="Total")
        cell.font = total_font
        cell.border = thin_border

        cell = ws.cell(row=current_row, column=2, value=totales['primas'])
        cell.font = total_font
        cell.border = thin_border
        cell.number_format = '#,##0,'
        cell.alignment = center_alignment

        cell = ws.cell(row=current_row, column=3, value=totales['variacion'])
        cell.font = total_font
        cell.number_format = '0.00'
        cell.alignment = center_alignment

        ws.cell(row=current_row, column=4, value="")
        ws.cell(row=current_row, column=5, value="")

        current_row += 3

    column_widths = [35, 17, 11, 11.5, 11.5]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def crear_hoja_detalle(ws, df):
    ws.sheet_view.showGridLines = False

    _, header_font, normal_font, total_font, center_alignment, center_vertical_alignment, thin_border = _estilos()

    headers = ["Subramo", "ENTIDAD", "Producción", "Var %", "Primas anterior"]
    columns_map = ['subramo_denominacion', 'nombre_corto', 'primas', 'variacion', 'primas_anterior']

    current_row = 1

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_vertical_alignment

    ws.row_dimensions[current_row].height = 39
    current_row += 1

    datos_ordenados = df.sort_values(['subramo_denominacion', 'nombre_corto'])

    for _, row_data in datos_ordenados.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=row_data[csv_col])
            cell.font = normal_font
            cell.border = thin_border

            if col_idx in [3, 5]:  # Producción y Primas anterior
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
            elif col_idx == 4:  # Var %
                cell.number_format = '0.00'
                cell.alignment = center_alignment

        current_row += 1

    # Total general
    totales = calcular_totales_subramo(df)

    cell = ws.cell(row=current_row, column=1, value="TOTAL GENERAL")
    cell.font = total_font
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=2, value="")
    cell.font = total_font
    cell.border = thin_border

    for col_idx, valor in zip([3, 4, 5], [totales['primas'], totales['variacion'], totales['primas_anterior']]):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = center_alignment
        if col_idx in [3, 5]:
            cell.number_format = '#,##0,'
        else:
            cell.number_format = '0.00'

    column_widths = [35, 35, 17, 11, 17]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def calcular_totales_subramo(data: pd.DataFrame) -> dict:
    primas = data['primas'].sum()
    primas_anterior = data['primas_anterior'].sum()
    variacion = ((primas / primas_anterior) - 1) * 100 if primas_anterior > 0 else 0.0
    return {'primas': primas, 'variacion': variacion, 'primas_anterior': primas_anterior}


def generate_apertura_subramo_comparativo_excel(period: str) -> str:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    csv_dir = os.path.join(base_dir, "ending_files", period)
    period_dir = os.path.join(base_dir, "excel_final_files", period)
    os.makedirs(period_dir, exist_ok=True)

    csv_path = os.path.join(csv_dir, f"{period}_apertura_por_subramo_comparativo.csv")
    output_path = os.path.join(period_dir, f"{period}_apertura_por_subramo_comparativo.xlsx")

    create_excel_apertura_subramo_comparativo(csv_path, output_path, period)
    return output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Genera Excel para apertura por subramo con comparativo')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')

    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO)

    try:
        excel_path = generate_apertura_subramo_comparativo_excel(args.period)
        print(f"✅ Excel generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_apertura_por_subramo_comparativo.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
