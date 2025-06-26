import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import os
import logging

def create_excel_ganaron_perdieron_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'ganaron_perdieron' con todas las hojas por tipo.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Convertir porcentajes de coma a punto para que Excel los reconozca como números
    pct_columns = ['pct_rt', 'pct_rf', 'pct_result']
    for col in pct_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').astype(float)
    
    # Obtener tipos únicos
    tipos = df['tipo_cia'].unique()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear una hoja para cada tipo
    for tipo in tipos:
        tipo_data = df[df['tipo_cia'] == tipo].copy()
        ws = wb.create_sheet(title=tipo)
        
        crear_hoja_tipo(ws, tipo_data, tipo)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel completo generado con {len(tipos)} hojas en: {output_path}")

def crear_hoja_tipo(ws, tipo_data, tipo_nombre):
    """Crea una hoja individual para un tipo de compañía"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    title_font = Font(name="Arial", size=11, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers de columnas
    headers = [
        "ENTIDAD", "Resultados técnicos", "% s/primas devengadas", 
        "Resultado financiero", "% s/primas devengadas", 
        "Resultado de operaciones extraordinarias", "Impuesto a las Ganancias", 
        "Resultado del ejercicio", "% s/primas devengadas"
    ]
    
    # Mapeo de columnas CSV a posiciones Excel
    columns_map = [
        'nombre_corto', 'resultado_tecnico', 'pct_rt', 
        'resultado_financiero', 'pct_rf', 'resultado_operaciones',
        'impuesto_ganancias', 'resultado', 'pct_result'
    ]
    
    # Separar en ganaron y perdieron y ordenar alfabéticamente
    ganaron = tipo_data[tipo_data['resultado'] > 0].copy().sort_values('nombre_corto')
    perdieron = tipo_data[tipo_data['resultado'] < 0].copy().sort_values('nombre_corto')
    
    current_row = 1
    
    # === ENCABEZADO FIJO ===
    ws.cell(row=current_row, column=1, value="RESULTADOS").font = title_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="JULIO/24 - MARZO/25, EN MILES DE PESOS").font = normal_font
    current_row += 2
    
    # === CUADRO 1: LAS QUE GANARON ===
    ws.cell(row=current_row, column=1, value="LAS QUE GANARON").font = title_font
    current_row += 2
    
    if not ganaron.empty:
        current_row = crear_cuadro_datos(ws, ganaron, headers, columns_map, current_row, 
                                       header_font, normal_font, total_font, center_alignment, center_vertical_alignment, thin_border)
        current_row += 2
    else:
        # Si no hay datos, solo mostrar headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_vertical_alignment
        ws.row_dimensions[current_row].height = 43
        current_row += 3
    
    # === CUADRO 2: LAS QUE PERDIERON ===
    ws.cell(row=current_row, column=1, value="LAS QUE PERDIERON").font = title_font
    current_row += 2
    
    if not perdieron.empty:
        current_row = crear_cuadro_datos(ws, perdieron, headers, columns_map, current_row,
                                       header_font, normal_font, total_font, center_alignment, center_vertical_alignment, thin_border)
        current_row += 2
    else:
        # Si no hay datos, solo mostrar headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_vertical_alignment
        ws.row_dimensions[current_row].height = 43
        current_row += 3
    
    # === TOTAL DEL MERCADO (para este tipo) ===
    totales_tipo = calcular_totales(tipo_data)
    
    cell = ws.cell(row=current_row, column=1, value=f"TOTAL {tipo_nombre.upper()}")
    cell.font = total_font
    cell.border = thin_border
    
    valores_mercado = [
        totales_tipo['resultado_tecnico'],
        round(totales_tipo['resultado_tecnico'] / totales_tipo['primas_devengadas'] * 100, 1),
        totales_tipo['resultado_financiero'],
        round(totales_tipo['resultado_financiero'] / totales_tipo['primas_devengadas'] * 100, 1),
        totales_tipo['resultado_operaciones'],
        totales_tipo['impuesto_ganancias'],
        totales_tipo['resultado'],
        round(totales_tipo['resultado'] / totales_tipo['primas_devengadas'] * 100, 1)
    ]
    
    for col_idx, valor in enumerate(valores_mercado, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        if col_idx not in [3, 5, 9]:
            cell.number_format = '#,##0,'
        else:  # Porcentajes con 1 decimal
            cell.number_format = '0.0'
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    column_widths = [35, 15, 12, 15, 12, 18, 15, 15, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

def crear_cuadro_datos(ws, data, headers, columns_map, start_row, header_font, normal_font, total_font, center_alignment, center_vertical_alignment, thin_border):
    """Crea un cuadro de datos (ganaron o perdieron) y retorna la siguiente fila disponible"""
    current_row = start_row
    
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx > 1:
            cell.alignment = center_vertical_alignment  # Centrado horizontal y vertical
        # La columna ENTIDAD (col_idx=1) mantiene el formato original
    
    # Ajustar altura de la fila de headers
    ws.row_dimensions[current_row].height = 43
    current_row += 1
    
    # Datos
    for _, row_data in data.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            value = row_data[csv_col]
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear números
            if col_idx > 1 and col_idx not in [3, 5, 9]:  # Columnas B,D,F,G,H
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
            elif col_idx in [3, 5, 9]:  # Porcentajes con 1 decimal
                cell.number_format = '0.0'
                cell.alignment = center_alignment
        current_row += 1
    
    # Total del cuadro
    totales = calcular_totales(data)
    
    cell = ws.cell(row=current_row, column=1, value="TOTAL")
    cell.font = total_font
    cell.border = thin_border
    
    valores_totales = [
        totales['resultado_tecnico'],
        round(totales['resultado_tecnico'] / totales['primas_devengadas'] * 100, 1),
        totales['resultado_financiero'],
        round(totales['resultado_financiero'] / totales['primas_devengadas'] * 100, 1),
        totales['resultado_operaciones'],
        totales['impuesto_ganancias'],
        totales['resultado'],
        round(totales['resultado'] / totales['primas_devengadas'] * 100, 1)
    ]
    
    for col_idx, valor in enumerate(valores_totales, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        if col_idx not in [3, 5, 9]:
            cell.number_format = '#,##0,'
        else:  # Porcentajes con 1 decimal
            cell.number_format = '0.0'
        cell.alignment = center_alignment
    
    return current_row + 1

def calcular_totales(data: pd.DataFrame) -> dict:
    """Calcula totales para un DataFrame de datos financieros"""
    return {
        'resultado_tecnico': data['resultado_tecnico'].sum(),
        'resultado_financiero': data['resultado_financiero'].sum(),
        'resultado_operaciones': data['resultado_operaciones'].sum(),
        'impuesto_ganancias': data['impuesto_ganancias'].sum(),
        'resultado': data['resultado'].sum(),
        'primas_devengadas': data['primas_devengadas'].sum()
    }

def generate_ganaron_perdieron_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel completo de ganaron-perdieron.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_ganaron_perdieron.csv")
    output_path = os.path.join(period_dir, f"{period}_ganaron_perdieron.xlsx")
    
    create_excel_ganaron_perdieron_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel completo para ganaron-perdieron')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_ganaron_perdieron_excel(args.period)
        print(f"✅ Excel completo generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_ganaron_perdieron.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")