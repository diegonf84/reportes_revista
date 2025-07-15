import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

def create_excel_sueldos_gastos_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'sueldos_y_gastos'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Obtener tipos únicos
    tipos_disponibles = sorted(df['tipo_cia'].unique())
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear una hoja para cada tipo
    for tipo in tipos_disponibles:
        tipo_data = df[df['tipo_cia'] == tipo].copy()
        ws = wb.create_sheet(title=tipo)
        crear_hoja_tipo(ws, tipo_data)
    
    # Crear hoja base_detalle
    ws_detalle = wb.create_sheet(title="base_detalle")
    crear_hoja_detalle(ws_detalle, df)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel generado con {len(tipos_disponibles)} + 1 hojas en: {output_path}")

def crear_hoja_tipo(ws, tipo_data):
    """Crea una hoja individual para un tipo de compañía con cálculos de porcentajes"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Empezar en fila 2 (primera fila vacía)
    current_row = 2
    
    # Headers
    headers = ["ENTIDAD", "% sueldos y cs", "% total gs prod", "% total gs"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_vertical_alignment
    
    # Ajustar altura de la fila de headers
    ws.row_dimensions[current_row].height = 39
    current_row += 1
    
    # Datos ordenados por primas devengadas descendente
    datos_ordenados = tipo_data.sort_values('total_primas_devengadas', ascending=False)
    
    for _, row_data in datos_ordenados.iterrows():
        # Calcular porcentajes
        pct_sueldos = (row_data['total_sueldos'] / row_data['total_primas_devengadas']) * 100 if row_data['total_primas_devengadas'] != 0 else 0
        pct_gs_prod = (row_data['total_gs_prod'] / row_data['total_primas_devengadas']) * 100 if row_data['total_primas_devengadas'] != 0 else 0
        pct_total_gs = (row_data['total_gs'] / row_data['total_primas_devengadas']) * 100 if row_data['total_primas_devengadas'] != 0 else 0
        
        # ENTIDAD
        cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
        cell_entidad.font = normal_font
        cell_entidad.border = thin_border
        
        # % SUELDOS Y CS
        cell_sueldos = ws.cell(row=current_row, column=2, value=pct_sueldos)
        cell_sueldos.font = normal_font
        cell_sueldos.border = thin_border
        cell_sueldos.number_format = '0.00'
        cell_sueldos.alignment = center_alignment
        
        # % TOTAL GS PROD
        cell_gs_prod = ws.cell(row=current_row, column=3, value=pct_gs_prod)
        cell_gs_prod.font = normal_font
        cell_gs_prod.border = thin_border
        cell_gs_prod.number_format = '0.00'
        cell_gs_prod.alignment = center_alignment
        
        # % TOTAL GS
        cell_total_gs = ws.cell(row=current_row, column=4, value=pct_total_gs)
        cell_total_gs.font = normal_font
        cell_total_gs.border = thin_border
        cell_total_gs.number_format = '0.00'
        cell_total_gs.alignment = center_alignment
        
        current_row += 1
    
    # Fila de totales
    totales = calcular_totales_tipo(datos_ordenados)
    
    # TOTAL
    cell_total = ws.cell(row=current_row, column=1, value="TOTAL")
    cell_total.font = total_font
    cell_total.border = thin_border
    
    # Calcular porcentajes totales
    pct_sueldos_total = (totales['total_sueldos'] / totales['total_primas_devengadas']) * 100 if totales['total_primas_devengadas'] != 0 else 0
    pct_gs_prod_total = (totales['total_gs_prod'] / totales['total_primas_devengadas']) * 100 if totales['total_primas_devengadas'] != 0 else 0
    pct_total_gs_total = (totales['total_gs'] / totales['total_primas_devengadas']) * 100 if totales['total_primas_devengadas'] != 0 else 0
    
    valores_totales = [pct_sueldos_total, pct_gs_prod_total, pct_total_gs_total]
    
    for col_idx, valor in enumerate(valores_totales, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        cell.number_format = '0.00'
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11

def crear_hoja_detalle(ws, df):
    """Crea la hoja 'base_detalle' con todos los valores sin calcular porcentajes"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Empezar en fila 2 (primera fila vacía)
    current_row = 2
    
    # Headers
    headers = ["TIPO", "ENTIDAD", "PRIMAS DEVENGADAS", "SUELDOS Y CARGAS SOCIALES", 
               "TOTAL GASTOS PRODUCCION", "TOTAL GASTOS", "GASTOS A/C REASEGURO"]
    columns_map = ['tipo_cia', 'nombre_corto', 'total_primas_devengadas', 'total_sueldos', 
                   'total_gs_prod', 'total_gs', 'gs_reaseguro']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_vertical_alignment
    
    # Ajustar altura de la fila de headers
    ws.row_dimensions[current_row].height = 39
    current_row += 1
    
    # Datos ordenados por tipo y entidad
    datos_ordenados = df.sort_values(['tipo_cia', 'nombre_corto'])
    
    for _, row_data in datos_ordenados.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            value = row_data[csv_col]
            
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear números (columnas 3-7 son numéricas)
            if col_idx >= 3:
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
        
        current_row += 1
    
    # Total general
    totales_generales = calcular_totales_generales(df)
    
    cell_total_tipo = ws.cell(row=current_row, column=1, value="TOTAL GENERAL")
    cell_total_tipo.font = total_font
    cell_total_tipo.border = thin_border
    
    cell_total_entidad = ws.cell(row=current_row, column=2, value="")
    cell_total_entidad.font = total_font
    cell_total_entidad.border = thin_border
    
    valores_totales = [
        totales_generales['total_primas_devengadas'],
        totales_generales['total_sueldos'],
        totales_generales['total_gs_prod'],
        totales_generales['total_gs'],
        totales_generales['gs_reaseguro']
    ]
    
    for col_idx, valor in enumerate(valores_totales, 3):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        cell.number_format = '#,##0,'
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    column_widths = [15, 40, 16, 16, 16, 16, 16]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def calcular_totales_tipo(data: pd.DataFrame) -> dict:
    """Calcula totales para un tipo específico"""
    return {
        'total_primas_devengadas': data['total_primas_devengadas'].sum(),
        'total_sueldos': data['total_sueldos'].sum(),
        'total_gs_prod': data['total_gs_prod'].sum(),
        'total_gs': data['total_gs'].sum(),
        'gs_reaseguro': data['gs_reaseguro'].sum()
    }

def calcular_totales_generales(data: pd.DataFrame) -> dict:
    """Calcula totales generales para toda la base"""
    return calcular_totales_tipo(data)

def generate_sueldos_gastos_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel de sueldos y gastos.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_sueldos_y_gastos.csv")
    output_path = os.path.join(period_dir, f"{period}_sueldos_y_gastos.xlsx")
    
    create_excel_sueldos_gastos_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel para sueldos y gastos')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_sueldos_gastos_excel(args.period)
        print(f"✅ Excel generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_sueldos_y_gastos.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")