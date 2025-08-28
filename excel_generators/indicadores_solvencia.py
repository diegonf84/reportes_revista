import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging
import argparse

def create_excel_indicadores_solvencia(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'indicadores_solvencia'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path)
    
    # No filtrar por tipo - incluir todos los tipos de compañía
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # === HOJA 1: INDICADORES SOLVENCIA ===
    ws_principal = wb.create_sheet(title="Indicadores Solvencia")
    crear_hoja_principal_solvencia(ws_principal, df)
    
    # === HOJA 2: BASE DETALLE ===
    ws_detalle = wb.create_sheet(title="base_detalle")
    crear_hoja_detalle_solvencia(ws_detalle, df)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel indicadores solvencia generado con hojas: Indicadores Solvencia, base_detalle en: {output_path}")

def crear_hoja_principal_solvencia(ws, df):
    """Crea la hoja principal 'Indicadores Solvencia' con ratios calculados agrupados por tipo"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    footnote_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Empezar en fila 2
    current_row = 2
    
    # Headers de columnas (reutilizable para cada sección)
    headers = ["ENTIDAD", "Disp+Inv / Ds.c/aseg %", "Disp+Inv+Inm / Ds.c/aseg %"]
    
    # Lista global para rastrear empresas con deudas = 0 (para footnote)
    empresas_sin_deudas = []
    
    # Agrupar por tipo de compañía
    tipos = sorted(df['tipo_cia'].unique())
    
    for tipo_idx, tipo in enumerate(tipos):
        # Título del tipo (bold + underlined)
        cell_titulo = ws.cell(row=current_row, column=1, value=tipo.upper())
        cell_titulo.font = title_font
        current_row += 1
        
        # Línea en blanco después del título
        current_row += 1
        
        # Headers para esta sección
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_vertical_alignment
        
        ws.row_dimensions[current_row].height = 27
        current_row += 1
        
        # Datos del tipo actual, ordenados alfabéticamente
        datos_tipo = df[df['tipo_cia'] == tipo].sort_values('nombre_corto')
        
        for _, row_data in datos_tipo.iterrows():
            # ENTIDAD en columna A
            cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # Calcular ratios o mostrar (*) si deudas = 0
            disponibilidades = row_data['disponibilidades']
            inversiones = row_data['inversiones']
            inmuebles_inversion = row_data['inmuebles_inversion']
            deudas_con_asegurados = row_data['deudas_con_asegurados']
            
            if deudas_con_asegurados == 0:
                # Mostrar (*) cuando no hay deudas
                empresas_sin_deudas.append(row_data['nombre_corto'])
                
                # Ratio 1: Disp+Inv / Ds.c/aseg %
                cell_ratio1 = ws.cell(row=current_row, column=2, value="(*)")
                cell_ratio1.font = normal_font
                cell_ratio1.border = thin_border
                cell_ratio1.alignment = center_alignment
                
                # Ratio 2: Disp+Inv+Inm / Ds.c/aseg %
                cell_ratio2 = ws.cell(row=current_row, column=3, value="(*)")
                cell_ratio2.font = normal_font
                cell_ratio2.border = thin_border
                cell_ratio2.alignment = center_alignment
            else:
                # Calcular ratios normalmente
                ratio1 = ((disponibilidades + inversiones) / deudas_con_asegurados) * 100
                ratio2 = ((disponibilidades + inversiones + inmuebles_inversion) / deudas_con_asegurados) * 100
                
                # Ratio 1: Disp+Inv / Ds.c/aseg %
                cell_ratio1 = ws.cell(row=current_row, column=2, value=ratio1)
                cell_ratio1.font = normal_font
                cell_ratio1.border = thin_border
                cell_ratio1.number_format = '#,##0.00'
                cell_ratio1.alignment = center_alignment
                
                # Ratio 2: Disp+Inv+Inm / Ds.c/aseg %
                cell_ratio2 = ws.cell(row=current_row, column=3, value=ratio2)
                cell_ratio2.font = normal_font
                cell_ratio2.border = thin_border
                cell_ratio2.number_format = '#,##0.00'
                cell_ratio2.alignment = center_alignment
            
            current_row += 1
        
        # Totales del tipo
        totales_tipo = calcular_totales_solvencia(datos_tipo)
        
        cell_total = ws.cell(row=current_row, column=1, value=f"Total {tipo}")
        cell_total.font = total_font
        cell_total.border = thin_border
        
        # Total ratio 1
        cell_total_ratio1 = ws.cell(row=current_row, column=2, value=totales_tipo['ratio1'])
        cell_total_ratio1.font = total_font
        cell_total_ratio1.border = thin_border
        cell_total_ratio1.number_format = '#,##0.00'
        cell_total_ratio1.alignment = center_alignment
        
        # Total ratio 2  
        cell_total_ratio2 = ws.cell(row=current_row, column=3, value=totales_tipo['ratio2'])
        cell_total_ratio2.font = total_font
        cell_total_ratio2.border = thin_border
        cell_total_ratio2.number_format = '#,##0.00'
        cell_total_ratio2.alignment = center_alignment
        
        current_row += 1
        
        # Dos líneas en blanco después de cada tipo (excepto el último)
        if tipo_idx < len(tipos) - 1:
            current_row += 2
    
    # Agregar footnote si hay empresas sin deudas (2 filas después del último total)
    if empresas_sin_deudas:
        current_row += 2
        cell_footnote = ws.cell(row=current_row, column=1, value="(*) No registra deudas con asegurados")
        cell_footnote.font = footnote_font
    
    # Ajustar ancho de columnas según especificaciones
    column_widths = [37, 14, 14]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def crear_hoja_detalle_solvencia(ws, df):
    """Crea la hoja 'base_detalle' con toda la información de solvencia agrupada por tipo"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers (reutilizable para cada sección)
    headers = [
        "Tipo", "ENTIDAD", "Inmuebles Inv.", "Inversiones", "Disponibilidades",
        "Deudas c/Aseg.", "Ratio 1 %", "Ratio 2 %"
    ]
    columns_map = [
        'tipo_cia', 'nombre_corto', 'inmuebles_inversion', 'inversiones', 
        'disponibilidades', 'deudas_con_asegurados'
    ]
    
    # Empezar en fila 2
    current_row = 2
    
    # Agrupar por tipo de compañía
    tipos = sorted(df['tipo_cia'].unique())
    
    for tipo_idx, tipo in enumerate(tipos):
        # Título del tipo (bold + underlined)
        cell_titulo = ws.cell(row=current_row, column=1, value=tipo.upper())
        cell_titulo.font = title_font
        current_row += 1
        
        # Línea en blanco después del título
        current_row += 1
        
        # Headers para esta sección
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_vertical_alignment
        
        ws.row_dimensions[current_row].height = 27
        current_row += 1
        
        # Datos del tipo actual, ordenados alfabéticamente
        datos_tipo = df[df['tipo_cia'] == tipo].sort_values('nombre_corto')
        
        for _, row_data in datos_tipo.iterrows():
            # Columnas básicas
            for col_idx, csv_col in enumerate(columns_map, 1):
                value = row_data[csv_col]
                
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                
                # Formatear según tipo de columna
                if col_idx in [3, 4, 5, 6]:  # Montos
                    cell.number_format = '#,##0,'
                    cell.alignment = center_alignment
            
            # Calcular y agregar ratios
            disponibilidades = row_data['disponibilidades']
            inversiones = row_data['inversiones']
            inmuebles_inversion = row_data['inmuebles_inversion']
            deudas_con_asegurados = row_data['deudas_con_asegurados']
            
            if deudas_con_asegurados == 0:
                # Ratio 1
                cell_ratio1 = ws.cell(row=current_row, column=7, value="(*)")
                cell_ratio1.font = normal_font
                cell_ratio1.border = thin_border
                cell_ratio1.alignment = center_alignment
                
                # Ratio 2
                cell_ratio2 = ws.cell(row=current_row, column=8, value="(*)")
                cell_ratio2.font = normal_font
                cell_ratio2.border = thin_border
                cell_ratio2.alignment = center_alignment
            else:
                ratio1 = ((disponibilidades + inversiones) / deudas_con_asegurados) * 100
                ratio2 = ((disponibilidades + inversiones + inmuebles_inversion) / deudas_con_asegurados) * 100
                
                # Ratio 1
                cell_ratio1 = ws.cell(row=current_row, column=7, value=ratio1)
                cell_ratio1.font = normal_font
                cell_ratio1.border = thin_border
                cell_ratio1.number_format = '#,##0.00'
                cell_ratio1.alignment = center_alignment
                
                # Ratio 2
                cell_ratio2 = ws.cell(row=current_row, column=8, value=ratio2)
                cell_ratio2.font = normal_font
                cell_ratio2.border = thin_border
                cell_ratio2.number_format = '#,##0.00'
                cell_ratio2.alignment = center_alignment
            
            current_row += 1
        
        # Total del tipo
        totales_tipo = calcular_totales_solvencia(datos_tipo)
        
        cell_total_tipo = ws.cell(row=current_row, column=1, value=f"TOTAL {tipo.upper()}")
        cell_total_tipo.font = total_font
        cell_total_tipo.border = thin_border
        
        cell_total_entidad = ws.cell(row=current_row, column=2, value="")
        cell_total_entidad.font = total_font
        cell_total_entidad.border = thin_border
        
        # Totales de montos para este tipo
        valores_totales_tipo = [
            totales_tipo['inmuebles_inversion'],
            totales_tipo['inversiones'],
            totales_tipo['disponibilidades'],
            totales_tipo['deudas_con_asegurados'],
            totales_tipo['ratio1'],
            totales_tipo['ratio2']
        ]
        
        for col_idx, valor in enumerate(valores_totales_tipo, 3):
            cell = ws.cell(row=current_row, column=col_idx, value=valor)
            cell.font = total_font
            cell.border = thin_border
            
            if col_idx in [3, 4, 5, 6]:  # Montos
                cell.number_format = '#,##0,'
            elif col_idx in [7, 8]:  # Ratios
                cell.number_format = '#,##0.00'
            
            cell.alignment = center_alignment
        
        current_row += 1
        
        # Dos líneas en blanco después de cada tipo (excepto el último)
        if tipo_idx < len(tipos) - 1:
            current_row += 2
    
    # Ajustar anchos de columnas
    column_widths = [15, 37, 14, 14, 14, 14, 11, 14]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def calcular_totales_solvencia(data: pd.DataFrame) -> dict:
    """Calcula totales generales para los indicadores de solvencia"""
    total_inmuebles = data['inmuebles_inversion'].sum()
    total_inversiones = data['inversiones'].sum()
    total_disponibilidades = data['disponibilidades'].sum()
    total_deudas = data['deudas_con_asegurados'].sum()
    
    # Calcular ratios totales (agregados, no promedio)
    if total_deudas > 0:
        ratio1 = ((total_disponibilidades + total_inversiones) / total_deudas) * 100
        ratio2 = ((total_disponibilidades + total_inversiones + total_inmuebles) / total_deudas) * 100
    else:
        ratio1 = 0.0
        ratio2 = 0.0
    
    return {
        'inmuebles_inversion': total_inmuebles,
        'inversiones': total_inversiones,
        'disponibilidades': total_disponibilidades,
        'deudas_con_asegurados': total_deudas,
        'ratio1': ratio1,
        'ratio2': ratio2
    }

def generate_indicadores_solvencia_excel(period: str, csv_dir: str = None) -> str:
    """
    Función de conveniencia para generar el Excel de indicadores de solvencia.
    
    Args:
        period: Período (ej: "202501")
        csv_dir: Directorio donde está el CSV (default: ending_files/{period}/)
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    if csv_dir is None:
        csv_dir = os.path.join(base_dir, "ending_files", period)
    
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_indicadores_solvencia.csv")
    output_path = os.path.join(period_dir, f"{period}_indicadores_solvencia.xlsx")
    
    create_excel_indicadores_solvencia(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Genera Excel formateado para indicadores de solvencia')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    parser.add_argument('--csv_dir', default=None, help='Directorio donde está el CSV')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    try:
        excel_path = generate_indicadores_solvencia_excel(
            period=args.period,
            csv_dir=args.csv_dir
        )
        print(f"✅ Excel generado exitosamente: {excel_path}")
    except FileNotFoundError as e:
        print(f"❌ Error: No se encontró el archivo CSV esperado")
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        expected_csv_path = f"{base_dir}/ending_files/{args.period}" if args.csv_dir is None else args.csv_dir
        print(f"   Buscando CSV en: {expected_csv_path}/{args.period}_indicadores_solvencia.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")