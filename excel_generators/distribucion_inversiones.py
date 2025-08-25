import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging
import argparse

def create_excel_distribucion_inversiones(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'distribucion_inversiones'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # === HOJA 1: DISTRIBUCION INVERSIONES ===
    ws_principal = wb.create_sheet(title="Distribucion Inversiones")
    crear_hoja_principal_inversiones(ws_principal, df)
    
    # === HOJA 2: BASE DETALLE ===
    ws_detalle = wb.create_sheet(title="base_detalle")
    crear_hoja_detalle_inversiones(ws_detalle, df)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel distribución inversiones generado con hojas: Distribucion Inversiones, base_detalle en: {output_path}")

def crear_hoja_principal_inversiones(ws, df):
    """Crea la hoja principal 'Distribucion Inversiones' con solo porcentajes"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Procesar cada tipo de compañía
    for tipo_cia in df['tipo_cia'].unique():
        
        # Filtrar datos para este tipo
        data_tipo = df[df['tipo_cia'] == tipo_cia].copy()
        
        # Título del tipo (ej: "ART", "Generales") en columna A
        cell_tipo = ws.cell(row=current_row, column=1, value=tipo_cia)
        cell_tipo.font = title_font
        current_row += 1
        
        # Headers de columnas, comenzando en columna B
        headers = ["ENTIDAD", "Total de inversiones", "% de inmuebles de inversion", "% de inversiones liquidas"]
        for col_idx, header in enumerate(headers, 2):  # Empezar en columna 2 (B)
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_vertical_alignment
        
        ws.row_dimensions[current_row].height = 49
        current_row += 1
        
        # Datos de las compañías, ordenados alfabéticamente
        data_tipo_ordenado = data_tipo.sort_values('nombre_corto')
        for _, row_data in data_tipo_ordenado.iterrows():
            # Nombre de la entidad en columna B
            cell_entidad = ws.cell(row=current_row, column=2, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # Total de inversiones en columna C
            cell_total_inv = ws.cell(row=current_row, column=3, value=row_data['total_inv'])
            cell_total_inv.font = normal_font
            cell_total_inv.border = thin_border
            cell_total_inv.number_format = '#,##0,'
            cell_total_inv.alignment = center_alignment
            
            # Porcentajes comenzando en columna D
            percentage_columns = ['pct_inmuebles_inversion', 'pct_inversiones_liquidas']
            for col_idx, csv_col in enumerate(percentage_columns, 4):
                value = row_data[csv_col]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.number_format = '0.00'
                cell.alignment = center_alignment
            
            current_row += 1
        
        # Fila de totales
        cell_total = ws.cell(row=current_row, column=2, value="Total")
        cell_total.font = total_font
        cell_total.border = thin_border
        
        # Total de inversiones
        total_inv = data_tipo['total_inv'].sum()
        cell_total_inv = ws.cell(row=current_row, column=3, value=total_inv)
        cell_total_inv.font = total_font
        cell_total_inv.border = thin_border
        cell_total_inv.number_format = '#,##0,'
        cell_total_inv.alignment = center_alignment
        
        # Calcular y escribir totales de porcentajes (promedios ponderados)
        if total_inv > 0:
            pct_values = [
                (data_tipo['total_inv_inm'].sum() / total_inv) * 100,
                (data_tipo['total_inv_liq'].sum() / total_inv) * 100
            ]
        else:
            pct_values = [0, 0]
        
        for col_idx, pct_value in enumerate(pct_values, 4):
            cell = ws.cell(row=current_row, column=col_idx, value=pct_value)
            cell.font = total_font
            cell.border = thin_border
            cell.number_format = '0.00'
            cell.alignment = center_alignment
        
        current_row += 2  # Espacio entre tipos
    
    # Ajustar ancho de columnas
    column_widths = [15, 36, 18, 14, 14]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def crear_hoja_detalle_inversiones(ws, df):
    """Crea la hoja 'base_detalle' con toda la información de inversiones"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Tipo", "ENTIDAD", "Total Inversiones", "Inmuebles Inv.", "Inversiones Liq.", 
        "% Inmuebles Inv.", "% Inversiones Liq."
    ]
    columns_map = [
        'tipo_cia', 'nombre_corto', 'total_inv', 'total_inv_inm', 
        'total_inv_liq', 'pct_inmuebles_inversion', 'pct_inversiones_liquidas'
    ]
    
    current_row = 1
    
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_vertical_alignment
    
    ws.row_dimensions[current_row].height = 34
    current_row += 1
    
    # Datos ordenados por tipo y entidad
    datos_ordenados = df.sort_values(['tipo_cia', 'nombre_corto'])
    
    for _, row_data in datos_ordenados.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            value = row_data[csv_col]
            
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear según tipo de columna
            if col_idx in [3, 4, 5]:  # Montos
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
            elif col_idx in [6, 7]:  # Porcentajes
                cell.number_format = '0.00'
                cell.alignment = center_alignment
        
        current_row += 1
    
    # Total general
    totales_generales = calcular_totales_generales_inversiones(df)
    
    cell_total_tipo = ws.cell(row=current_row, column=1, value="TOTAL GENERAL")
    cell_total_tipo.font = total_font
    cell_total_tipo.border = thin_border
    
    cell_total_entidad = ws.cell(row=current_row, column=2, value="")
    cell_total_entidad.font = total_font
    cell_total_entidad.border = thin_border
    
    valores_totales = [
        totales_generales['total_inv'],
        totales_generales['total_inv_inm'],
        totales_generales['total_inv_liq'],
        totales_generales['pct_inmuebles_inversion'],
        totales_generales['pct_inversiones_liquidas']
    ]
    
    for col_idx, valor in enumerate(valores_totales, 3):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        
        if col_idx in [3, 4, 5]:  # Montos
            cell.number_format = '#,##0,'
        elif col_idx in [6, 7]:  # Porcentajes
            cell.number_format = '0.00'
        
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    column_widths = [12, 36, 16, 14, 14, 14, 14]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def calcular_totales_generales_inversiones(data: pd.DataFrame) -> dict:
    """Calcula totales generales para toda la base de inversiones"""
    total_inv = data['total_inv'].sum()
    total_inv_inm = data['total_inv_inm'].sum()
    total_inv_liq = data['total_inv_liq'].sum()
    
    # Calcular porcentajes totales
    if total_inv > 0:
        pct_inmuebles = (total_inv_inm / total_inv) * 100
        pct_liquidas = (total_inv_liq / total_inv) * 100
    else:
        pct_inmuebles = pct_liquidas = 0
    
    return {
        'total_inv': total_inv,
        'total_inv_inm': total_inv_inm,
        'total_inv_liq': total_inv_liq,
        'pct_inmuebles_inversion': pct_inmuebles,
        'pct_inversiones_liquidas': pct_liquidas
    }

def generate_distribucion_inversiones_excel(period: str, csv_dir: str = None) -> str:
    """
    Función de conveniencia para generar el Excel de distribución de inversiones.
    
    Args:
        period: Período (ej: "202501")
        csv_dir: Directorio donde está el CSV (default: ending_files/{period}/)
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    if csv_dir is None:
        csv_dir = os.path.join(base_dir, "ending_files", period)
    
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_distribucion_inversiones.csv")
    output_path = os.path.join(period_dir, f"{period}_distribucion_inversiones.xlsx")
    
    create_excel_distribucion_inversiones(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Genera Excel formateado para distribución de inversiones')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    parser.add_argument('--csv_dir', default=None, help='Directorio donde está el CSV')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    try:
        excel_path = generate_distribucion_inversiones_excel(
            period=args.period,
            csv_dir=args.csv_dir
        )
        print(f"✅ Excel generado exitosamente: {excel_path}")
    except FileNotFoundError as e:
        print(f"❌ Error: No se encontró el archivo CSV esperado")
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        expected_csv_path = f"{base_dir}/ending_files/{args.period}" if args.csv_dir is None else args.csv_dir
        print(f"   Buscando CSV en: {expected_csv_path}/{args.period}_distribucion_inversiones.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")