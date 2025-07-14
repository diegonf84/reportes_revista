import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

def create_excel_ranking_ramo_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'ranking_comparativo_por_ramo'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Convertir variación de coma a punto
    df['variacion'] = df['variacion'].astype(str).str.replace(',', '.').astype(float)
    
    # Obtener todos los ramos disponibles (sin filtros como en tipos)
    ramos_disponibles = sorted(df['ramo_denomincion'].unique())
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # === HOJA 1: RANKING ===
    ws_ranking = wb.create_sheet(title="Ranking")
    crear_hoja_ranking_ramos(ws_ranking, df, ramos_disponibles)
    
    # === HOJA 2: BASE DETALLE ===
    ws_detalle = wb.create_sheet(title="base_detalle")
    crear_hoja_detalle_ramos(ws_detalle, df)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel generado con hojas: Ranking, base_detalle en: {output_path}")

def crear_hoja_ranking_ramos(ws, df, ramos_disponibles):
    """Crea la hoja principal 'Ranking' con cuadros por ramo"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    title_font = Font(name="Arial", size=10, bold=True, underline="single")
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Procesar cada ramo
    for ramo in ramos_disponibles:
        ramo_data = df[df['ramo_denomincion'] == ramo].copy()
        
        # Título del ramo
        cell_ramo = ws.cell(row=current_row, column=1, value=ramo)
        cell_ramo.font = title_font
        current_row += 2
        
        # Headers
        headers = ["ENTIDAD", "Primas emitidas", "Var %"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            if col_idx != 3:  # No aplicar borde a columna C (Var %)
                cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_vertical_alignment
            else:
                cell.alignment = center_vertical_alignment
        
        ws.row_dimensions[current_row].height = 39
        current_row += 1
        
        # Datos ordenados alfabéticamente
        datos_ordenados = ramo_data.sort_values('nombre_corto')
        
        for _, row_data in datos_ordenados.iterrows():
            # ENTIDAD
            cell_entidad = ws.cell(row=current_row, column=1, value=row_data['nombre_corto'])
            cell_entidad.font = normal_font
            cell_entidad.border = thin_border
            
            # PRIMAS EMITIDAS
            cell_primas = ws.cell(row=current_row, column=2, value=row_data['primas_emitidas'])
            cell_primas.font = normal_font
            cell_primas.border = thin_border
            cell_primas.number_format = '#,##0,'
            cell_primas.alignment = center_alignment
            
            # VAR %
            cell_var = ws.cell(row=current_row, column=3, value=row_data['variacion'])
            cell_var.font = normal_font
            # No aplicar borde a columna C (Var %)
            cell_var.number_format = '0.00'
            cell_var.alignment = center_alignment
            
            current_row += 1
        
        # Total del ramo
        totales = calcular_totales_ramo(ramo_data)
        
        cell_total = ws.cell(row=current_row, column=1, value="Total")
        cell_total.font = total_font
        cell_total.border = thin_border
        
        cell_total_primas = ws.cell(row=current_row, column=2, value=totales['primas_emitidas'])
        cell_total_primas.font = total_font
        cell_total_primas.border = thin_border
        cell_total_primas.number_format = '#,##0,'
        cell_total_primas.alignment = center_alignment
        
        cell_total_var = ws.cell(row=current_row, column=3, value=totales['variacion'])
        cell_total_var.font = total_font
        # No aplicar borde a columna C (Var %)
        cell_total_var.number_format = '0.00'
        cell_total_var.alignment = center_alignment
        
        current_row += 3  # Espacio entre ramos
    
    # Ajustar anchos de columnas
    column_widths = [36, 16, 11]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def crear_hoja_detalle_ramos(ws, df):
    """Crea la hoja 'base_detalle' con toda la información por ramos"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Ramo", "ENTIDAD", "Primas emitidas", "Variación %", "Primas anterior"]
    columns_map = ['ramo_denomincion', 'nombre_corto', 'primas_emitidas', 'variacion', 'primas_anterior']
    
    current_row = 1
    
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_vertical_alignment
    
    ws.row_dimensions[current_row].height = 39
    current_row += 1
    
    # Datos ordenados por ramo y entidad
    datos_ordenados = df.sort_values(['ramo_denomincion', 'nombre_corto'])
    
    for _, row_data in datos_ordenados.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            value = row_data[csv_col]
            
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear según tipo de columna
            if col_idx in [3, 5]:  # Primas emitidas y anteriores
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
            elif col_idx == 4:  # Variación %
                cell.number_format = '0.00'
                cell.alignment = center_alignment
        
        current_row += 1
    
    # Total general
    totales_generales = calcular_totales_generales_ramos(df)
    
    cell_total = ws.cell(row=current_row, column=1, value="TOTAL GENERAL")
    cell_total.font = total_font
    cell_total.border = thin_border
    
    cell_total_entidad = ws.cell(row=current_row, column=2, value="")
    cell_total_entidad.font = total_font
    cell_total_entidad.border = thin_border
    
    valores_totales = [
        totales_generales['primas_emitidas'],
        totales_generales['variacion'],
        totales_generales['primas_anterior']
    ]
    
    for col_idx, valor in enumerate(valores_totales, 3):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        
        if col_idx in [3, 5]:  # Primas
            cell.number_format = '#,##0,'
        elif col_idx == 4:  # Variación
            cell.number_format = '0.00'
        
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    column_widths = [25, 36, 16, 11, 16]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def calcular_totales_ramo(data: pd.DataFrame) -> dict:
    """Calcula totales para un ramo específico"""
    primas_actual = data['primas_emitidas'].sum()
    primas_anterior = data['primas_anterior'].sum()
    
    # Calcular variación: ((actual/anterior)-1)*100
    if primas_anterior > 0:
        variacion = ((primas_actual / primas_anterior) - 1) * 100
    else:
        variacion = 0.0
    
    return {
        'primas_emitidas': primas_actual,
        'variacion': variacion,
        'primas_anterior': primas_anterior
    }

def calcular_totales_generales_ramos(data: pd.DataFrame) -> dict:
    """Calcula totales generales para toda la base de ramos"""
    return calcular_totales_ramo(data)

def generate_ranking_ramo_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel de ranking comparativo por ramo.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_ranking_comparativo_por_ramo.csv")
    output_path = os.path.join(period_dir, f"{period}_ranking_comparativo_por_ramo.xlsx")
    
    create_excel_ranking_ramo_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel para ranking comparativo por ramo')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_ranking_ramo_excel(args.period)
        print(f"✅ Excel generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_ranking_comparativo_por_ramo.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")