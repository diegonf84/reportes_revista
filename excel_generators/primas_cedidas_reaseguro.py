import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

def create_excel_primas_cedidas_completo(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel completo para el reporte 'primas_cedidas_reaseguro'.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Convertir porcentajes de coma a punto
    pct_columns = ['pct_cesion', 'pct_ret']
    for col in pct_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').astype(float)
    
    # Filtrar tipos (excluir Retiro) - usando nombres del CSV
    tipos_incluir = ['ART', 'Generales', 'M.T.P.P.', 'Vida']  # M.T.P.P. como aparece en CSV
    tipos_disponibles = [tipo for tipo in df['tipo_cia'].unique() if tipo in tipos_incluir]
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear una hoja para cada tipo
    for tipo in tipos_disponibles:
        tipo_data = df[df['tipo_cia'] == tipo].copy()
        # Nombre de hoja: convertir M.T.P.P. -> MTTP para Excel
        nombre_hoja = "MTTP" if tipo == "M.T.P.P." else tipo
        ws = wb.create_sheet(title=nombre_hoja)
        crear_hoja_tipo(ws, tipo_data)
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel completo generado con {len(tipos_disponibles)} hojas en: {output_path}")

def crear_hoja_tipo(ws, tipo_data):
    """Crea una hoja individual para un tipo de compañía"""
    
    # Configurar hoja
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    center_vertical_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers de columnas
    headers = [
        "ENTIDAD", "Producción", "Primas cedidas al reaseguro", "% / total",
        "Primas netas de reaseguro", "% / total"
    ]
    
    # Mapeo de columnas CSV a posiciones Excel
    columns_map = [
        'nombre_corto', 'primas_emitidas', 'primas_cedidas', 'pct_cesion',
        'primas_retenidas', 'pct_ret'
    ]
    
    # Ordenar alfabéticamente por entidad
    datos_ordenados = tipo_data.sort_values('nombre_corto')
    
    current_row = 2  # Empezar en fila 2 (sin título general)
    
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        if col_idx > 1:  # Centrar todas excepto ENTIDAD (que también se centra ahora)
            cell.alignment = center_vertical_alignment
        else:  # ENTIDAD también centrada vertical y horizontalmente
            cell.alignment = center_vertical_alignment
    
    # Ajustar altura de la fila de headers
    ws.row_dimensions[current_row].height = 39
    current_row += 1
    
    # Datos de las entidades
    for _, row_data in datos_ordenados.iterrows():
        for col_idx, csv_col in enumerate(columns_map, 1):
            value = row_data[csv_col]
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatear según tipo de columna
            if col_idx == 1:  # ENTIDAD - sin formato especial
                pass
            elif col_idx in [2, 3, 5]:  # Valores monetarios
                cell.number_format = '#,##0,'
                cell.alignment = center_alignment
            elif col_idx in [4, 6]:  # Porcentajes
                cell.number_format = '0.00'
                cell.alignment = center_alignment
        
        current_row += 1
    
    # Fila de totales
    totales = calcular_totales(datos_ordenados)
    
    # TOTAL
    cell_total = ws.cell(row=current_row, column=1, value="TOTAL")
    cell_total.font = total_font
    cell_total.border = thin_border
    
    valores_totales = [
        totales['primas_emitidas'],
        totales['primas_cedidas'],
        totales['pct_cesion'],
        totales['primas_retenidas'],
        totales['pct_ret']
    ]
    
    for col_idx, valor in enumerate(valores_totales, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        
        if col_idx in [2, 3, 5]:  # Valores monetarios
            cell.number_format = '#,##0,'
        elif col_idx in [4, 6]:  # Porcentajes
            cell.number_format = '0.00'
        
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas según especificaciones
    column_widths = [36, 16, 16, 11, 16, 11]  # A=36, B=16, C=16, D=11, E=16, F=11
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def calcular_totales(data: pd.DataFrame) -> dict:
    """Calcula totales y porcentajes para el resumen"""
    totales = {
        'primas_emitidas': data['primas_emitidas'].sum(),
        'primas_cedidas': data['primas_cedidas'].sum(),
        'primas_retenidas': data['primas_retenidas'].sum()
    }
    
    # Calcular porcentajes basados en totales
    if totales['primas_emitidas'] > 0:
        totales['pct_cesion'] = (totales['primas_cedidas'] / totales['primas_emitidas']) * 100
        totales['pct_ret'] = (totales['primas_retenidas'] / totales['primas_emitidas']) * 100
    else:
        totales['pct_cesion'] = 0.0
        totales['pct_ret'] = 0.0
    
    return totales

def generate_primas_cedidas_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel de primas cedidas.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_primas_cedidas_reaseguro.csv")
    output_path = os.path.join(period_dir, f"{period}_primas_cedidas_reaseguro.xlsx")
    
    create_excel_primas_cedidas_completo(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel para primas cedidas reaseguro')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_primas_cedidas_excel(args.period)
        print(f"✅ Excel generado: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_primas_cedidas_reaseguro.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")