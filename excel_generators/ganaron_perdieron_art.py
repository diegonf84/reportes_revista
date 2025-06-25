import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import os
import logging

def create_excel_ganaron_perdieron_art(csv_path: str, output_path: str, period: str) -> None:
    """
    Genera archivo Excel formateado para el reporte 'ganaron_perdieron' - solo ART para prueba.
    
    Args:
        csv_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el Excel
        period: Período del reporte (ej: "202501")
    """
    
    # Leer CSV
    df = pd.read_csv(csv_path, sep=';')
    
    # Filtrar solo ART para prueba
    art_data = df[df['tipo_cia'] == 'ART'].copy()
    
    # Separar en ganaron y perdieron
    ganaron = art_data[art_data['resultado'] > 0].copy()
    perdieron = art_data[art_data['resultado'] < 0].copy()
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ART"
    ws.sheet_view.showGridLines = False
    
    # Definir estilos
    title_font = Font(name="Arial", size=11, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers de columnas
    headers = [
        "ENTIDAD", "Resultados técnicos", "% s/primas devengadas", 
        "Resultado financiero", "% s/primas devengadas", 
        "Resultado de operaciones extraordinarias", "Impuesto a las Ganancias", 
        "Resultado del ejercicio", "% s/primas devengadas"
    ]
    
    # Mapeo de columnas CSV a posiciones Excel
    columns_map = [
        'nombre_corto', 'resultado_tecnico', 'pct_rt', 
        'resultado_financiero', 'pct_rf', 'resultado_operaciones',
        'impuesto_ganancias', 'resultado', 'pct_result'
    ]
    
    current_row = 1
    
    # === ENCABEZADO FIJO ===
    ws.cell(row=current_row, column=1, value="RESULTADOS").font = title_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="JULIO/24 - MARZO/25, EN MILES DE PESOS").font = normal_font
    current_row += 2
    
    # === CUADRO 1: LAS QUE GANARON ===
    ws.cell(row=current_row, column=1, value="LAS QUE GANARON").font = title_font
    current_row += 2
    
    if not ganaron.empty:
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_alignment
        current_row += 1
        
        # Datos
        for _, row_data in ganaron.iterrows():
            for col_idx, csv_col in enumerate(columns_map, 1):
                value = row_data[csv_col]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                
                # Formatear números
                if col_idx > 1 and col_idx not in [3, 5, 9]:  # No formatear porcentajes
                    cell.number_format = '#,##0'
                    cell.alignment = center_alignment
                elif col_idx in [3, 5, 9]:  # Porcentajes
                    cell.alignment = center_alignment
            current_row += 1
        
        # Total cuadro ganaron
        totales_ganaron = calcular_totales(ganaron)
        
        cell = ws.cell(row=current_row, column=1, value="TOTAL")
        cell.font = total_font
        cell.border = thin_border
        
        # Valores totales
        valores_totales = [
            totales_ganaron['resultado_tecnico'],
            round(totales_ganaron['resultado_tecnico'] / totales_ganaron['primas_devengadas'] * 100, 1),
            totales_ganaron['resultado_financiero'],
            round(totales_ganaron['resultado_financiero'] / totales_ganaron['primas_devengadas'] * 100, 1),
            totales_ganaron['resultado_operaciones'],
            totales_ganaron['impuesto_ganancias'],
            totales_ganaron['resultado'],
            round(totales_ganaron['resultado'] / totales_ganaron['primas_devengadas'] * 100, 1)
        ]
        
        for col_idx, valor in enumerate(valores_totales, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=valor)
            cell.font = total_font
            cell.border = thin_border
            if col_idx not in [3, 5, 9]:  # No formatear porcentajes
                cell.number_format = '#,##0'
            cell.alignment = center_alignment
        
        current_row += 3
    
    # === CUADRO 2: LAS QUE PERDIERON ===
    ws.cell(row=current_row, column=1, value="LAS QUE PERDIERON").font = title_font
    current_row += 2
    
    if not perdieron.empty:
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_alignment
        current_row += 1
        
        # Datos
        for _, row_data in perdieron.iterrows():
            for col_idx, csv_col in enumerate(columns_map, 1):
                value = row_data[csv_col]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                
                if col_idx > 1 and col_idx not in [3, 5, 9]:
                    cell.number_format = '#,##0'
                    cell.alignment = center_alignment
                elif col_idx in [3, 5, 9]:
                    cell.alignment = center_alignment
            current_row += 1
        
        # Total cuadro perdieron
        totales_perdieron = calcular_totales(perdieron)
        
        cell = ws.cell(row=current_row, column=1, value="TOTAL")
        cell.font = total_font
        cell.border = thin_border
        
        valores_totales = [
            totales_perdieron['resultado_tecnico'],
            round(totales_perdieron['resultado_tecnico'] / totales_perdieron['primas_devengadas'] * 100, 1),
            totales_perdieron['resultado_financiero'],
            round(totales_perdieron['resultado_financiero'] / totales_perdieron['primas_devengadas'] * 100, 1),
            totales_perdieron['resultado_operaciones'],
            totales_perdieron['impuesto_ganancias'],
            totales_perdieron['resultado'],
            round(totales_perdieron['resultado'] / totales_perdieron['primas_devengadas'] * 100, 1)
        ]
        
        for col_idx, valor in enumerate(valores_totales, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=valor)
            cell.font = total_font
            cell.border = thin_border
            if col_idx not in [3, 5, 9]:
                cell.number_format = '#,##0'
            cell.alignment = center_alignment
        
        current_row += 3
    
    # === TOTAL DEL MERCADO ===
    totales_mercado = calcular_totales(art_data)
    
    cell = ws.cell(row=current_row, column=1, value="TOTAL DEL MERCADO")
    cell.font = total_font
    cell.border = thin_border
    
    valores_mercado = [
        totales_mercado['resultado_tecnico'],
        round(totales_mercado['resultado_tecnico'] / totales_mercado['primas_devengadas'] * 100, 1),
        totales_mercado['resultado_financiero'],
        round(totales_mercado['resultado_financiero'] / totales_mercado['primas_devengadas'] * 100, 1),
        totales_mercado['resultado_operaciones'],
        totales_mercado['impuesto_ganancias'],
        totales_mercado['resultado'],
        round(totales_mercado['resultado'] / totales_mercado['primas_devengadas'] * 100, 1)
    ]
    
    for col_idx, valor in enumerate(valores_mercado, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=valor)
        cell.font = total_font
        cell.border = thin_border
        if col_idx not in [3, 5, 9]:
            cell.number_format = '#,##0'
        cell.alignment = center_alignment
    
    # Ajustar anchos de columnas
    column_widths = [25, 15, 12, 15, 12, 18, 15, 15, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Guardar archivo
    wb.save(output_path)
    logging.info(f"Excel ART generado en: {output_path}")

def calcular_totales(data: pd.DataFrame) -> dict:
    """Calcula totales para un DataFrame de datos financieros"""
    return {
        'resultado_tecnico': data['resultado_tecnico'].sum(),
        'resultado_financiero': data['resultado_financiero'].sum(),
        'resultado_operaciones': data['resultado_operaciones'].sum(),
        'impuesto_ganancias': data['impuesto_ganancias'].sum(),
        'resultado': data['resultado'].sum(),
        'primas_devengadas': data['primas_devengadas'].sum()
    }

def generate_ganaron_perdieron_art_excel(period: str) -> str:
    """
    Función de conveniencia para generar el Excel de ganaron-perdieron ART.
    
    Args:
        period: Período (ej: "202501")
        
    Returns:
        Ruta del archivo Excel generado
    """
    # Obtener directorio base del proyecto (subir un nivel desde excel_generators)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    csv_dir = os.path.join(base_dir, "ending_files", period)
    output_dir = os.path.join(base_dir, "excel_final_files")
    period_dir = os.path.join(output_dir, period)
    os.makedirs(period_dir, exist_ok=True)
    
    csv_path = os.path.join(csv_dir, f"{period}_ganaron_perdieron.csv")
    output_path = os.path.join(period_dir, f"{period}_ganaron_perdieron_art_test.xlsx")
    
    create_excel_ganaron_perdieron_art(csv_path, output_path, period)
    return output_path

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera Excel ART para ganaron-perdieron')
    parser.add_argument('period', help='Período del reporte (ej: 202501)')
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    try:
        excel_path = generate_ganaron_perdieron_art_excel(args.period)
        print(f"✅ Excel generado exitosamente: {excel_path}")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo CSV esperado")
        print(f"   Buscando: ending_files/{args.period}/{args.period}_ganaron_perdieron.csv")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")